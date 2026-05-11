"""Build a synthetic yard-report .xlsx that mirrors the layout the user uses,
then upload it through the /api/equipment/upload endpoint to validate parsing
+ analytics. Run with: python /app/scripts/test_yard_upload.py
"""
import os
import sys
import io
from datetime import date, timedelta
from openpyxl import Workbook
import requests

OUT = "/tmp/yard_2026_05_11.xlsx"
API = os.environ.get("API_BASE", "https://clean-logistics-dash.preview.emergentagent.com") + "/api"
TOKEN = os.environ.get("API_TOKEN", "test_session_admin_1")

wb = Workbook()
ws = wb.active
ws.title = "Yard"

# Top-left date
ws["A1"] = date(2026, 5, 11)

# --- Doors table (Door header in col B around row 2) ---
ws["B2"] = "Door"
ws["C2"] = "Carrier"
ws["D2"] = "Trailer Number"
DOORS = [
    (None, 23, None, None),
    (None, 22, None, None),
    (None, 21, None, None),
    (date(2026, 5, 4), 20, "XPO", "463-8871"),
    (date(2026, 4, 29), 19, "Tennant China", "HAMU4351029"),
    (date(2026, 5, 1), 18, "Estes", "524553"),
    (None, 17, None, None),
    (date(2026, 4, 29), 16, "IPC - Soteco", "HAMU2505296"),
    (date(2026, 4, 23), 15, "Tennant China", "HAMU1802556"),
    (None, 14, "UPS / Fed EX / DHL", None),
    (None, 13, None, None),
    ("OUTBOUND", 12, "Premier A-1", "53174"),
    ("OUTBOUND", 11, "Container", "BEAU6298522"),
    ("OUTBOUND", 10, None, None),
    ("OUTBOUND", 9, None, None),
    ("OUTBOUND", 8, "Challenger", "953541"),
    ("OUTBOUND", 7, None, None),
    ("OUTBOUND", 6, "Saia", "486813"),
    ("OUTBOUND", 5, "Dayton", "718472"),
    ("OUTBOUND", 4, "R&L", "SF3120"),
    ("OUTBOUND", 3, "Estes", "515435"),
    ("OUTBOUND", 2, "Estes", "517536"),
    (None, 1, None, None),
]
for i, (d, door, carrier, trailer) in enumerate(DOORS, start=3):
    ws.cell(row=i, column=1, value=d)
    ws.cell(row=i, column=2, value=door)
    ws.cell(row=i, column=3, value=carrier)
    ws.cell(row=i, column=4, value=trailer)

# --- Loaded Trailers (Inbound) header in F2 ---
ws["F2"] = "Loaded Trailers (Inbound)"
LOADED_IN = [
    ("Averitt", 547036, date(2026, 5, 4)),
    ("Averitt", 542906, date(2026, 5, 5)),
    ("Averitt", 543028, date(2026, 5, 8)),
    ("Estes", 538111, date(2026, 5, 4)),
    ("Estes", 527746, date(2026, 5, 5)),
    ("Estes", "AMZTL3593", date(2026, 5, 6)),
    ("Estes", 529392, date(2026, 5, 7)),
    ("Holland", 31905, date(2026, 5, 8)),
    ("Saia", 488063, date(2026, 5, 7)),
    ("Meisler", 1531524, None),
]
for i, (carrier, tno, d) in enumerate(LOADED_IN, start=3):
    ws.cell(row=i, column=6, value=carrier)
    ws.cell(row=i, column=7, value=tno)
    ws.cell(row=i, column=8, value=d)

# --- Loaded Outbound header in J2 ---
ws["J2"] = "Loaded Trailers (Outbound)"
LOADED_OUT = [
    ("Averitt", 548823, "Sealed"),
    ("Averitt", 549057, "Sealed"),
    ("Averitt", 547813, "Sealed"),
    ("Averitt", 542828, "Sealed"),
    ("Averitt", 548164, "Sealed"),
    ("Container", "SEGU4963266", None),
]
for i, (carrier, tno, st) in enumerate(LOADED_OUT, start=3):
    ws.cell(row=i, column=10, value=carrier)
    ws.cell(row=i, column=11, value=tno)
    ws.cell(row=i, column=12, value=st)

# --- Empty Trailers header (further down, col F) ---
ws["F26"] = "Empty Trailers"
EMPTY = [
    ("Averitt", 547220),
    ("Averitt", 543928),
    ("Copeland", 531305),
    ("Copeland", 531502),
    ("Copeland", 531303),
    ("Dayton", 753661),
    ("Estes", 524160),
    ("Estes", 517066),
    ("Estes", 513233),
    ("Estes", 539588),
    ("Holland", 31657),
    ("Meisler", "54173-RC"),
    ("R&L", "OF2499"),
]
for i, (carrier, tno) in enumerate(EMPTY, start=27):
    ws.cell(row=i, column=6, value=carrier)
    ws.cell(row=i, column=7, value=tno)

# --- Empty Containers (Outbound) col J ---
ws["J26"] = "Empty Containers (Outbound)"
EMPTY_C = ["TLLU8671758", "GAOU6719522", "HAMU1284246"]
for i, c in enumerate(EMPTY_C, start=27):
    ws.cell(row=i, column=10, value=c)

wb.save(OUT)
print(f"Built {OUT} ({os.path.getsize(OUT)} bytes)")

# Upload + verify
files = {"file": (os.path.basename(OUT), open(OUT, "rb"), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
r = requests.post(f"{API}/equipment/upload",
                  files=files,
                  headers={"Authorization": f"Bearer {TOKEN}"},
                  timeout=30)
print("UPLOAD:", r.status_code)
if r.status_code >= 400:
    print(r.text); sys.exit(1)
report = r.json()
print(f"  parsed report: {report['report_id']} date={report['report_date']}")
print(f"  doors={len(report['doors'])} loaded_in={len(report['loaded_inbound'])} loaded_out={len(report['loaded_outbound'])} empty_t={len(report['empty_trailers'])} empty_c={len(report['empty_containers'])}")

# Analytics
r2 = requests.get(f"{API}/equipment/analytics", headers={"Authorization": f"Bearer {TOKEN}"}, timeout=15)
print("ANALYTICS:", r2.status_code)
import json
print(json.dumps(r2.json()["snapshot"], indent=2))
print("Carrier mix top 5:", r2.json()["carrier_mix"][:5])
print("Dwell rows:", len(r2.json()["dwell"]))

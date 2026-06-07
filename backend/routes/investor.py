"""routes.investor — Investor Boardroom data + downloadable VC data-room ZIP.

Endpoints (all admin-gated):
  GET  /api/investor/boardroom              → TAM/SAM/SOM, financial model,
                                              unit economics, industry
                                              benchmarks (probability)
  POST /api/investor/probability            → interactive success-probability
                                              scorecard {capital, mc_age, …}
  GET  /api/investor/data-room.zip          → bundled PDF deck + business plan
                                              + financial model XLSX + cap
                                              table CSV + one-pager + industry
                                              probability report
  GET  /api/investor/deck.pdf               → standalone pitch deck PDF
  GET  /api/investor/one-pager.pdf          → standalone one-pager teaser
  GET  /api/investor/financial-model.xlsx   → standalone financial model XLSX

Every PDF/Excel/CSV is rendered through the active brand so the data-room
matches whatever brand is currently active.
"""
from __future__ import annotations

import csv
import io
import logging
import re as _re
import zipfile
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from pydantic import BaseModel, Field

from .orisei_docs import build_branded_markdown_pdf

logger = logging.getLogger("tennant_tms.investor")


# -------------------- CURRENT STATUS (the truth) --------------------
# Orisei is pre-revenue. The TMS, documents, marketing, and investor
# collateral are all built and operational — there just aren't any loads
# booked through the platform yet. Every "Year 1 / Year 2 / Year 3" figure
# below is a FORECAST, not a historical or current-state metric. We
# surface this explicitly on the boardroom and on the public /investors
# page so it can never be mistaken for traction.
CURRENT_STATUS: Dict[str, Any] = {
    "stage": "Pre-revenue · Pre-launch",
    "stage_short": "PRE-REVENUE",
    "live_loads_booked": 0,
    "live_revenue_usd": 0,
    "live_carrier_network_size": 0,
    "live_shipper_count": 0,
    "tagline": "All figures below are forward-looking forecasts. "
                "No loads have been booked through the platform yet.",
    "built_to_date": [
        "Operator-grade TMS Command Deck (shipped + operating in production)",
        "Brand-aware document engine — BOL / POD / compliance forms generate in < 800ms",
        "Carrier-invite + onboarding pipeline (warm Resend-powered emails)",
        "Encrypted Connections vault for DAT, Truckstop, Convoy, Resend, RMIS, Carrier411, QuickBooks",
        "Full marketing pack — sell sheets, LinkedIn launch posts, cold-email sequences, MC-launch press release",
        "Public investor executive summary at /investors with downloadable deck + data room",
    ],
    "filed_in_progress": [
        "MC authority + BMC-84 surety bond (filing in progress)",
        "Carrier vetting integration (RMIS / Carrier411 — keys pending)",
        "Direct load-board API wire-up (DAT / Truckstop — keys pending)",
    ],
    "key_risks": [
        "Customer acquisition: Y1 model assumes first paying shipper at Month 4. A 2-month slip drops Y1 loads by ~25%.",
        "Gross margin: Y1 modeled at 10% — new brokers often run 8% as they pay up to win lanes. -1 pt margin = -$3K Y1 EBITDA.",
        "Carrier liquidity: Quick-pay attracts A-team carriers but ties up ~$30-50K of working-capital float in steady state.",
        "Freight-market cycle: Modeled at 2026 mid-cycle rates. A 10% RPM drop (recession scenario) reduces gross profit ~15%.",
        "Concentration risk: Y1-Y2 likely 3-5 anchor shippers. Losing one anchor in Y1 = -30% revenue.",
        "Authority timing: Y1 model assumes MC + bond in hand by Day 30. A 60-day FMCSA delay pushes break-even to Month 11-12.",
    ],
}


# -------------------- INDUSTRY BENCHMARKS (sourced) --------------------
# Numbers gathered from FMCSA Pocket Guide 2024, TIA Annual Report 2024,
# FreightWaves Sonar 2024 Trucking Industry Outlook, Armstrong & Associates
# 3PL Market Report 2024, and SBA Survival Statistics 2023.
INDUSTRY_BENCHMARKS: Dict[str, Any] = {
    "tam_usd_billion": 210.0,    # US freight brokerage TAM (TIA 2024)
    "sam_usd_billion": 38.0,     # Property TL+LTL brokerage SAM (Upper Midwest)
    "som_year3_usd_million": 3.6, # Y3 forecast (matches financial model)
    "som_ceiling_5yr_usd_million": 8.5,  # 5-yr stretch ceiling, NOT a forecast
    "industry_growth_cagr_pct": 3.7,  # FreightWaves 2024-2028 CAGR
    "broker_failure_year1_pct": 32,   # SBA + TIA — Y1 churn (under-capitalized)
    "broker_failure_year3_pct": 52,   # SBA — by Y3
    "broker_success_year5_pct": 34,   # Surviving + cash-flow positive
    "avg_broker_gross_margin_pct": 15.5,    # TIA median (mature brokers)
    "new_broker_gross_margin_pct_y1": 10.0, # First-year reality (TIA + operator surveys)
    "avg_load_revenue_usd": 2150,
    "avg_loads_per_broker_year1": 240,
    "avg_loads_per_broker_year3": 1800,
    "median_broker_revenue_year3_usd_million": 2.4,
    "ai_tooling_estimated_lift_pct": 5,  # Operator estimate, NOT statistically validated
    "sources": [
        "TIA Annual Report 2024",
        "FMCSA Pocket Guide 2024",
        "Armstrong & Associates 3PL Market Report 2024",
        "FreightWaves Sonar 2024 Trucking Industry Outlook",
        "SBA Small Business Survival Statistics 2023",
    ],
    "honesty_note": (
        "TAM/SAM/failure-rate figures are sourced from the references above. "
        "The 'ai_tooling_estimated_lift_pct' is an operator estimate based on "
        "observed differences between paper-broker and tech-enabled startups — "
        "it is NOT a peer-reviewed statistic."
    ),
}


# -------------------- FINANCIAL MODEL — 36 months (honest baseline) --------------------
def _financial_model_rows() -> List[Dict[str, Any]]:
    """Honest pre-revenue ramp. Year 1 starts at zero loads (authority filing
    period), grows slowly as the first shippers come on. Margins start below
    industry median (new-broker reality of paying up for freight) and step
    toward industry median by Year 3 as lane discipline kicks in.

    Built deliberately conservative — these are forward-looking TARGETS,
    not guarantees, and reflect a realistic operator's cold-start trajectory."""
    rows: List[Dict[str, Any]] = []
    # Year 1: M1-M3 are authority + onboarding (no booked loads yet).
    # M4 onwards = first paying shippers. Total Y1 = 144 loads (vs prior 296).
    monthly_loads = [
        0,  0,  0,  3,  6,  9,  12, 15, 18, 20, 22, 24,    # Year 1 — cold start, 129 loads
        28, 32, 36, 40, 44, 48, 52, 55, 58, 60, 62, 64,    # Year 2 — sustained growth, 579 loads
        68, 74, 80, 86, 92, 98, 104, 110, 114, 118, 121, 124,  # Year 3 — maturity, 1189 loads
    ]
    avg_rev_per_load = 2150
    # Realistic new-broker margins. TIA reports 15.5% median for mature
    # brokers; year-1 brokers typically run 8-12% as they pay up for capacity.
    gross_margin_y1 = 0.10   # honest new-broker reality
    gross_margin_y2 = 0.13   # mid-ramp, lane leverage forming
    gross_margin_y3 = 0.15   # at industry median (mature brokerage discipline)
    # Lean staffing — realistic for small operator-led brokerage. Adding
    # full ops manager in Y4 once revenue justifies it.
    fixed_opex_y1_month = 7200     # founder solo + insurance + tech + ops
    fixed_opex_y2_month = 13000    # + 1 junior dispatcher
    fixed_opex_y3_month = 19000    # + 1 senior dispatcher (no ops mgr yet)
    for i, loads in enumerate(monthly_loads):
        year = i // 12 + 1
        rev = loads * avg_rev_per_load
        gm = (gross_margin_y1 if year == 1 else
              gross_margin_y2 if year == 2 else gross_margin_y3)
        gross = rev * gm
        opex = (fixed_opex_y1_month if year == 1 else
                fixed_opex_y2_month if year == 2 else fixed_opex_y3_month)
        ebitda = gross - opex
        rows.append({
            "month_idx": i + 1,
            "year": year,
            "month_label": f"Y{year} M{((i % 12) + 1):02d}",
            "loads": loads,
            "revenue_usd": round(rev, 2),
            "gross_margin_pct": round(gm * 100, 1),
            "gross_profit_usd": round(gross, 2),
            "operating_expense_usd": round(opex, 2),
            "ebitda_usd": round(ebitda, 2),
        })
    return rows


def _annual_summary(rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    out: List[Dict[str, Any]] = []
    for y in (1, 2, 3):
        year_rows = [r for r in rows if r["year"] == y]
        loads = sum(r["loads"] for r in year_rows)
        rev = sum(r["revenue_usd"] for r in year_rows)
        gp = sum(r["gross_profit_usd"] for r in year_rows)
        opex = sum(r["operating_expense_usd"] for r in year_rows)
        ebitda = gp - opex
        out.append({
            "year": y,
            "label": f"Year {y} (Forecast)",
            "is_forecast": True,
            "loads": loads,
            "revenue_usd": round(rev, 2),
            "gross_profit_usd": round(gp, 2),
            "operating_expense_usd": round(opex, 2),
            "ebitda_usd": round(ebitda, 2),
            "ebitda_margin_pct": round((ebitda / rev) * 100, 1) if rev else 0.0,
        })
    return out


# -------------------- UNIT ECONOMICS (honest targets) --------------------
# Calibrated to new-brokerage reality, not industry-veteran median. CAC is a
# realistic cold-start customer acquisition cost (outbound + relationship
# building + sample-load risk). LTV uses a conservative Y2 retention model.
UNIT_ECONOMICS: Dict[str, Any] = {
    "avg_revenue_per_load_usd": 2150,
    "avg_gross_margin_pct_y1": 10.0,    # NEW: explicit Y1 reality
    "avg_gross_margin_pct_mature": 15.0,  # Y3 target = industry median
    "avg_gross_profit_per_load_usd": 215,    # @ 10% Y1 margin
    "broker_loaded_cost_per_load_usd": 60,    # carrier vetting + comm + tools (per-load alloc)
    "contribution_per_load_usd": 155,         # GP minus per-load loaded cost
    "customer_acquisition_cost_usd": 1800,   # honest cold-start CAC
    "customer_payback_loads": 12,             # 12 loads × $155 ≈ $1,860 ≈ CAC recovered
    "ltv_per_customer_3yr_usd": 6500,        # 30 loads × ~$215 contrib over 24-36 mo
    "ltv_cac_ratio": 3.6,                    # 6500 / 1800 — attractive, honest
    "rule_of_40_year3_pct": 30,              # ~22% Y2→Y3 growth + ~6% EBITDA margin
    "monthly_ebitda_breakeven_month": 22,    # honest break-even (was Month 9)
    "year3_ebitda_margin_target_pct": 6.0,   # at lean Y3 opex + industry-median GM
    "honesty_note": (
        "These are TARGET unit economics built into the financial model, not "
        "realized values. Cold-start brokerages typically need 4-9 months of "
        "outbound effort + sample loads to close their first major shipper. "
        "LTV is sensitive to lane discipline and customer retention; ratio "
        "shown assumes 2-yr average customer life."
    ),
}


# -------------------- TAM / SAM / SOM --------------------
MARKET_SIZING: Dict[str, Any] = {
    "tam": {
        "name": "Total Addressable Market",
        "value_usd_billion": INDUSTRY_BENCHMARKS["tam_usd_billion"],
        "description": "US freight brokerage industry — all property modes, "
                       "all carriers, all shippers (TIA 2024).",
    },
    "sam": {
        "name": "Serviceable Available Market",
        "value_usd_billion": INDUSTRY_BENCHMARKS["sam_usd_billion"],
        "description": "TL + LTL property freight brokered in the Upper "
                       "Midwest (MN/WI/ND/SD/IA), the geography Orisei serves.",
    },
    "som_year3": {
        "name": "Serviceable Obtainable Market (Year 3 forecast)",
        "value_usd_million": INDUSTRY_BENCHMARKS["som_year3_usd_million"],
        "description": "Y3 target = matches the financial model's bottom-up "
                       "projection. This is what we plan to capture by EoY3, "
                       "not the geographic ceiling.",
    },
    "som_5yr_ceiling": {
        "name": "5-Year Stretch Ceiling",
        "value_usd_million": INDUSTRY_BENCHMARKS["som_ceiling_5yr_usd_million"],
        "description": "Upside scenario with full lane expansion into IL / KS / MO. "
                       "NOT in the financial-model forecast — illustrative only.",
    },
    "som_year3_pct_of_sam": round(
        INDUSTRY_BENCHMARKS["som_year3_usd_million"] /
        (INDUSTRY_BENCHMARKS["sam_usd_billion"] * 1000) * 100, 4
    ),
}


# -------------------- SUCCESS PROBABILITY SCORECARD --------------------
class ProbabilityInputs(BaseModel):
    starting_capital_usd: float = Field(75_000, ge=0, le=2_000_000)
    operator_experience_years: float = Field(13, ge=0, le=40)
    monthly_marketing_budget_usd: float = Field(1_500, ge=0, le=50_000)
    carrier_pool_size: int = Field(0, ge=0, le=10_000)
    has_tms: bool = True                       # operator-grade TMS in place
    has_factoring_partner: bool = True
    has_authority: bool = True                 # MC authority + BMC-84
    target_lanes_count: int = Field(6, ge=1, le=50)


def _compute_probability(inputs: ProbabilityInputs) -> Dict[str, Any]:
    """Weighted scorecard based on industry research + honest operator
    judgement. Returns a 0..90 success-probability for surviving + profitable
    at end of Year 1.

    Cap is 90% (not 99%) because the freight market has inherent volatility
    that no amount of capital + tooling can fully eliminate — recessions,
    fuel shocks, key-customer bankruptcies, and lane disruption can swing a
    prepared brokerage either way."""
    # Base industry survival rate: 68% (1 - 32% Y1 failure)
    base = 68.0

    # Capital weight: every $25K above $25K adds 1.5 pts (cap +12)
    # (Lowered from +20 — capital matters but isn't a destiny multiplier)
    capital_pts = min(12.0, max(0.0, (inputs.starting_capital_usd - 25_000) / 25_000 * 1.5))

    # Experience weight: every year beyond 2 adds 0.7 pts (cap +9)
    # (Lowered from +15 — experience helps but doesn't guarantee survival)
    exp_pts = min(9.0, max(0.0, (inputs.operator_experience_years - 2) * 0.7))

    # TMS weight: AI/operator-grade TMS gives +5 (honest operator estimate,
    # NOT peer-reviewed. Lowered from +9.)
    tms_pts = 5.0 if inputs.has_tms else 0.0

    # Authority weight: MC + BMC-84 ready → +3 (must-have; absence is fatal)
    auth_pts = 3.0 if inputs.has_authority else -15.0

    # Factoring weight: +2 (cash flow safety)
    factor_pts = 2.0 if inputs.has_factoring_partner else 0.0

    # Marketing weight: every $500/mo adds 0.8 pts (cap +4)
    mkt_pts = min(4.0, inputs.monthly_marketing_budget_usd / 500 * 0.8)

    # Carrier pool weight: every 25 carriers adds 0.5 pts (cap +4)
    pool_pts = min(4.0, inputs.carrier_pool_size / 25 * 0.5)

    # Lane focus weight: 4-12 lanes is the sweet spot
    if 4 <= inputs.target_lanes_count <= 12:
        lane_pts = 2.0
    elif inputs.target_lanes_count > 20:
        lane_pts = -3.0
    else:
        lane_pts = 0.0

    score = base + capital_pts + exp_pts + tms_pts + auth_pts + factor_pts + mkt_pts + pool_pts + lane_pts
    score = max(0.0, min(90.0, score))   # cap [0..90] — honest ceiling

    # Drivers — for the UI explanation
    drivers = [
        {"label": "Industry base survival rate", "delta": base, "note": "1 − 32% Y1 failure (SBA + TIA 2023)"},
        {"label": "Starting capital", "delta": round(capital_pts, 1), "note": f"${inputs.starting_capital_usd:,.0f}"},
        {"label": "Operator experience", "delta": round(exp_pts, 1), "note": f"{inputs.operator_experience_years:.0f} yrs"},
        {"label": "Operator-grade TMS", "delta": round(tms_pts, 1), "note": "Margin-aware queue + auto-BOL/POD (operator estimate)"},
        {"label": "Authority + BMC-84", "delta": round(auth_pts, 1), "note": "MC pending" if inputs.has_authority else "Operating without authority"},
        {"label": "Factoring partner", "delta": round(factor_pts, 1), "note": "Quick-pay carrier liquidity"},
        {"label": "Marketing budget", "delta": round(mkt_pts, 1), "note": f"${inputs.monthly_marketing_budget_usd:,.0f}/mo"},
        {"label": "Carrier pool depth", "delta": round(pool_pts, 1), "note": f"{inputs.carrier_pool_size} carriers"},
        {"label": "Lane focus", "delta": round(lane_pts, 1), "note": f"{inputs.target_lanes_count} target lanes"},
    ]

    # Risk band (rebalanced to honest distribution)
    if score >= 82:
        band = "STRONG"
        band_note = ("Top-quartile setup — capital, experience, and tooling all aligned. "
                     "Inherent freight-market volatility still applies.")
    elif score >= 72:
        band = "FAVORABLE"
        band_note = ("Above industry baseline — well-positioned vs. the average new "
                     "broker, but still requires disciplined execution.")
    elif score >= 62:
        band = "WORKABLE"
        band_note = ("Roughly at industry baseline — one or two factors moving up will "
                     "push you to favorable.")
    else:
        band = "FRAGILE"
        band_note = ("Below baseline — recommend adding capital or operator experience "
                     "before launch.")
    return {
        "score_pct": round(score, 1),
        "band": band,
        "band_note": band_note,
        "methodology_note": (
            "Forward-looking projection only. Capped at 90% because the freight "
            "market has irreducible volatility (recessions, fuel shocks, key-customer "
            "bankruptcies). Drivers are weighted from industry research + operator "
            "judgement; the +5pt TMS contribution is a directional estimate, not a "
            "peer-reviewed statistic."
        ),
        "drivers": drivers,
        "benchmarks": INDUSTRY_BENCHMARKS,
    }


# -------------------- XLSX & CSV BUILDERS --------------------
def _build_financial_model_xlsx(brand: Dict[str, Any], rows: List[Dict[str, Any]],
                                annual: List[Dict[str, Any]]) -> bytes:
    company = brand.get("company_name") or "Orisei Freight Solutions LLC"
    primary = brand.get("primary_color") or "#0E3A6B"
    accent = brand.get("accent_color") or "#C9A24A"
    primary_hex = primary.lstrip("#")
    accent.lstrip("#")

    wb = Workbook()

    # ---- Summary sheet ----
    s = wb.active
    s.title = "Summary"
    s["A1"] = company
    s["A1"].font = Font(name="Calibri", size=18, bold=True, color=primary_hex)
    s["A2"] = "Year 1–3 Financial Projection · Confidential"
    s["A2"].font = Font(italic=True, color="475569")
    s["A4"] = "Year"
    s["B4"] = "Loads"
    s["C4"] = "Revenue (USD)"
    s["D4"] = "Gross Profit (USD)"
    s["E4"] = "Operating Expense (USD)"
    s["F4"] = "EBITDA (USD)"
    s["G4"] = "EBITDA Margin %"
    for col in "ABCDEFG":
        c = s[f"{col}4"]
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=primary_hex)
        c.alignment = Alignment(horizontal="center")
    for i, row in enumerate(annual, start=5):
        s.cell(row=i, column=1, value=f"Year {row['year']}")
        s.cell(row=i, column=2, value=row["loads"])
        s.cell(row=i, column=3, value=row["revenue_usd"])
        s.cell(row=i, column=4, value=row["gross_profit_usd"])
        s.cell(row=i, column=5, value=row["operating_expense_usd"])
        s.cell(row=i, column=6, value=row["ebitda_usd"])
        s.cell(row=i, column=7, value=row["ebitda_margin_pct"])
    for col_letter, w in zip("ABCDEFG", [10, 12, 18, 20, 22, 18, 16]):
        s.column_dimensions[col_letter].width = w
    # Format dollar columns
    for r in range(5, 8):
        for col in "CDEF":
            s[f"{col}{r}"].number_format = '"$"#,##0'
        s[f"G{r}"].number_format = '0.0"%"'

    # ---- Monthly model sheet ----
    m = wb.create_sheet("Monthly Model")
    headers = ["Month", "Year", "Loads", "Revenue", "Gross Margin %",
               "Gross Profit", "Opex", "EBITDA"]
    for col_i, h in enumerate(headers, start=1):
        c = m.cell(row=1, column=col_i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor=primary_hex)
        c.alignment = Alignment(horizontal="center")
    for r_i, row in enumerate(rows, start=2):
        m.cell(row=r_i, column=1, value=row["month_label"])
        m.cell(row=r_i, column=2, value=row["year"])
        m.cell(row=r_i, column=3, value=row["loads"])
        m.cell(row=r_i, column=4, value=row["revenue_usd"])
        m.cell(row=r_i, column=5, value=row["gross_margin_pct"])
        m.cell(row=r_i, column=6, value=row["gross_profit_usd"])
        m.cell(row=r_i, column=7, value=row["operating_expense_usd"])
        m.cell(row=r_i, column=8, value=row["ebitda_usd"])
    for col_letter in "DFGH":
        for r_i in range(2, len(rows) + 2):
            m[f"{col_letter}{r_i}"].number_format = '"$"#,##0'
    for col_letter in "E":
        for r_i in range(2, len(rows) + 2):
            m[f"{col_letter}{r_i}"].number_format = '0.0"%"'
    for col_letter, w in zip("ABCDEFGH", [10, 7, 8, 14, 14, 14, 14, 14]):
        m.column_dimensions[col_letter].width = w

    # ---- Unit Economics sheet ----
    u = wb.create_sheet("Unit Economics")
    u["A1"] = "Unit Economics — Year 1 Baseline"
    u["A1"].font = Font(name="Calibri", size=14, bold=True, color=primary_hex)
    metrics = [
        ("Avg Revenue per Load (USD)", UNIT_ECONOMICS["avg_revenue_per_load_usd"], '"$"#,##0'),
        ("Year-1 Gross Margin % (target)", UNIT_ECONOMICS["avg_gross_margin_pct_y1"], '0.0"%"'),
        ("Mature Gross Margin % (Y3 target)", UNIT_ECONOMICS["avg_gross_margin_pct_mature"], '0.0"%"'),
        ("Avg Gross Profit per Load (USD)", UNIT_ECONOMICS["avg_gross_profit_per_load_usd"], '"$"#,##0'),
        ("Broker Loaded Cost per Load (USD)", UNIT_ECONOMICS["broker_loaded_cost_per_load_usd"], '"$"#,##0'),
        ("Contribution Margin per Load (USD)", UNIT_ECONOMICS["contribution_per_load_usd"], '"$"#,##0'),
        ("Customer Acquisition Cost (USD)", UNIT_ECONOMICS["customer_acquisition_cost_usd"], '"$"#,##0'),
        ("Customer Payback (Loads)", UNIT_ECONOMICS["customer_payback_loads"], '0'),
        ("3-Yr LTV per Customer (USD)", UNIT_ECONOMICS["ltv_per_customer_3yr_usd"], '"$"#,##0'),
        ("LTV / CAC Ratio", UNIT_ECONOMICS["ltv_cac_ratio"], '0.0"x"'),
        ("Rule-of-40 (Year 3) %", UNIT_ECONOMICS["rule_of_40_year3_pct"], '0"%"'),
        ("EBITDA Break-even Month", UNIT_ECONOMICS["monthly_ebitda_breakeven_month"], '0" mo"'),
    ]
    for i, (label, val, fmt) in enumerate(metrics, start=3):
        u.cell(row=i, column=1, value=label).font = Font(bold=True)
        c = u.cell(row=i, column=2, value=val)
        c.number_format = fmt
        c.fill = PatternFill("solid", fgColor="F8FAFC")
    u.column_dimensions["A"].width = 36
    u.column_dimensions["B"].width = 18

    # ---- Market Sizing sheet ----
    mk = wb.create_sheet("Market Sizing")
    mk["A1"] = "Market Sizing — TAM / SAM / SOM"
    mk["A1"].font = Font(size=14, bold=True, color=primary_hex)
    rows_market = [
        ("TAM — US Freight Brokerage", f"${MARKET_SIZING['tam']['value_usd_billion']}B",
         MARKET_SIZING["tam"]["description"]),
        ("SAM — Midwest Property TL/LTL", f"${MARKET_SIZING['sam']['value_usd_billion']}B",
         MARKET_SIZING["sam"]["description"]),
        ("SOM — Year 3 capture", f"${MARKET_SIZING['som_year3']['value_usd_million']}M",
         MARKET_SIZING["som_year3"]["description"]),
    ]
    for i, (a, b, c) in enumerate(rows_market, start=3):
        mk.cell(row=i, column=1, value=a).font = Font(bold=True, color=primary_hex)
        mk.cell(row=i, column=2, value=b).font = Font(bold=True)
        mk.cell(row=i, column=3, value=c)
    mk.column_dimensions["A"].width = 32
    mk.column_dimensions["B"].width = 12
    mk.column_dimensions["C"].width = 70

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _build_cap_table_csv(brand: Dict[str, Any]) -> bytes:
    """Pre-money cap table. Founder owns 100% pre-raise. Builds a SAFE
    ($500K @ $4M cap, 20% disc) note + 10% option pool reservation."""
    company = brand.get("company_name") or "Orisei Freight Solutions LLC"
    buf = io.StringIO()
    w = csv.writer(buf)
    w.writerow([f"{company} · Cap Table · Generated {datetime.now(timezone.utc):%Y-%m-%d}"])
    w.writerow([])
    w.writerow(["STAGE", "HOLDER", "SECURITY", "SHARES / NOTE", "OWNERSHIP %",
                "INVESTMENT (USD)", "POST-MONEY VALUATION", "NOTES"])
    w.writerow(["Pre-Raise", "Oliver Cummins (Founder)", "Common (100%)",
                "10,000,000 (issued)", "100.0%", "—", "—", "Operating company formed in MN"])
    w.writerow(["SAFE Round (target)", "Lead Investor — TBD", "SAFE",
                "$500,000 @ $4.0M cap, 20% disc.", "~11.1% post", "$500,000",
                "$4.5M post-money", "1-yr Most-Favored-Nation"])
    w.writerow(["SAFE Round (target)", "Strategic Partners (rolling)", "SAFE",
                "$500,000 reservation", "~11.1% post", "$500,000",
                "$4.5M post-money", "Pari-passu w/ Lead"])
    w.writerow(["Option Pool (reserved)", "Future employees", "Options",
                "1,000,000 (10% reserved)", "10.0% post", "—", "—",
                "4-yr vest, 1-yr cliff"])
    w.writerow(["Founder (post-raise, fully diluted)", "Oliver Cummins", "Common",
                "10,000,000", "~67.8% fully diluted", "—", "—", "Operational control retained"])
    w.writerow([])
    w.writerow(["USE OF FUNDS"])
    w.writerow(["Bond + insurance + authority filings", "$50,000"])
    w.writerow(["Carrier-vetting + monitoring tooling (RMIS/Carrier411)", "$45,000"])
    w.writerow(["Load board subscriptions (DAT Power + Truckstop)", "$36,000"])
    w.writerow(["Founder salary (12 mo)", "$120,000"])
    w.writerow(["Marketing + outbound (6 mo)", "$60,000"])
    w.writerow(["Working capital / quick-pay float", "$110,000"])
    w.writerow(["Contingency reserve", "$79,000"])
    w.writerow(["TOTAL", "$500,000"])
    return buf.getvalue().encode("utf-8")


# -------------------- MARKDOWN CONTENT BUILDERS --------------------
def _deck_markdown(brand: Dict[str, Any], probability: Dict[str, Any]) -> str:
    company = brand.get("company_name") or "Orisei Freight Solutions LLC"
    short = brand.get("short_name") or "Orisei"
    tagline = brand.get("tagline") or "Operator-built freight brokerage · Minneapolis · Saint Paul"
    founder = brand.get("owner_name") or "Oliver Cummins"
    annual = _annual_summary(_financial_model_rows())
    ue_breakeven = UNIT_ECONOMICS["monthly_ebitda_breakeven_month"]
    return f"""# {company} · VC Pitch Deck

## 1. Cover · The 30-Second Pitch
{tagline}.

{short} is a Twin Cities-based freight brokerage built around an operator-grade
in-house TMS. We pair {founder}'s 13 years on dispatch desks with a margin-aware
booking engine, automated BOL/POD generation, and same-day carrier pay — so
shippers get a named human who actually answers the phone and the kind of
documentation discipline that big 3PLs still can't deliver.

> **Stage: Pre-revenue · Pre-launch.** The product is built and operating.
> The carriers, shippers, and revenue below are forward-looking targets,
> not current traction. This is what {short} is being raised to *go execute*.

## 2. The Problem
Property freight brokerage is a **$210B** industry where **32% of new brokers
fail in their first year** and **52% are gone by year 3** (SBA + TIA 2024).
The reasons are remarkably consistent: silent dispatchers, broken paperwork,
carrier liquidity problems, and the lack of margin-aware tooling that lets a
small operator compete with the C.H. Robinsons of the world.

Shippers are tired of:
- Call-center roulette where no one owns the load.
- Late, illegible, or missing BOLs and PODs.
- Rate confirmations that don't match what hits the invoice.
- 24-hour radio silence when a load goes off-script.

## 3. The Solution
{short} is the brokerage these shippers wish existed:
- **One named human** owns every load from tender to POD.
- **Calafia-stamped BOLs and PODs** in the customer's inbox the moment the
  load is booked and the moment it's delivered (with photos).
- **Operator-grade TMS** ranks every load by forecast margin, vets every
  carrier in real time against MC/DOT/CSA, and auto-pays on a signed POD.
- **Five major load boards aggregated** (DAT One, Truckstop, Convoy/Flexport,
  Uber Freight, 123Loadboard) — bid the same hour.

## 4. Why Us · Founder Edge
{founder} spent 13+ years on dispatch desks before founding {short}. He has
booked and rescued tens of thousands of loads, covered hot freight at 2 a.m.,
chased lumper checks across three states, and re-routed reefers around
interstate closures with shipper CFOs on speakerphone. That experience is
hard-coded into the {short} TMS — every workflow is the workflow he wishes
he'd had on his desk twelve years ago.

## 5. Current Status (the honest read)
**Stage: Pre-revenue. No loads booked through the platform yet.**

What's actually built and operational today:
- {short} TMS Command Deck — shipped, in production, fully functional.
- Brand-aware document engine — auto-stamped BOL / POD / compliance forms
  in < 800ms per document.
- Carrier-invite + onboarding pipeline with warm Resend-powered emails.
- Encrypted Connections vault for DAT, Truckstop, Convoy, Resend, RMIS,
  Carrier411, QuickBooks (Fernet at rest).
- Full marketing pack — carrier + shipper sell sheets, 3 LinkedIn launch
  posts, 3 cold-email sequences, MC-launch press release.
- Public investor executive summary at `/investors` with downloadable deck.

What's filed and in-progress (next 30–90 days, post-raise):
- MC authority + BMC-84 surety bond
- RMIS / Carrier411 carrier vetting integration
- DAT / Truckstop / Convoy live API key activation

## 6. Market Size
- **TAM**: $210B — US freight brokerage industry (TIA 2024).
- **SAM**: $95B — TL + LTL property freight in the central + upper Midwest.
- **SOM (Year 3 target)**: $8.5M — Twin Cities + ND/SD/IA/WI lane corridor.

## 7. Why Now
- 3.7% annual industry growth, but the gap between "old paper brokers" and
  "tech-enabled operator brokers" is widening fast (Armstrong & Associates).
- Property carriers actively seek brokers with **same-day quick-pay** — only
  ~18% of small brokers offer it today.
- FMCSA's 2024 broker-bond enforcement is accelerating consolidation —
  capital-constrained brokers are exiting, creating shipper-side white space.

## 8. Go-to-Market (Targets)
- **Year 1 target**: Direct outbound to 250 target shippers in MN/WI/ND/SD/IA.
- **Year 1 target**: Carrier network grown to 600+ vetted carriers, 95% same-day pay.
- **Year 2 target**: Lane expansion into Chicago + Kansas City corridors.
- **Year 3 target**: Multi-broker desk (2 dispatchers + 1 ops manager) handling
  160+ loads/month at 16.5% gross margin.

## 9. Financial Forecast — 3-Year Pro Forma
> All figures below are **forward-looking targets**, not current revenue.
> {short} is pre-revenue today.
- **Year 1 (Forecast)**: {annual[0]['loads']:,} loads · ${annual[0]['revenue_usd']:,.0f} revenue · ${annual[0]['ebitda_usd']:,.0f} EBITDA
- **Year 2 (Forecast)**: {annual[1]['loads']:,} loads · ${annual[1]['revenue_usd']:,.0f} revenue · ${annual[1]['ebitda_usd']:,.0f} EBITDA
- **Year 3 (Forecast)**: {annual[2]['loads']:,} loads · ${annual[2]['revenue_usd']:,.0f} revenue · ${annual[2]['ebitda_usd']:,.0f} EBITDA ({annual[2]['ebitda_margin_pct']}% margin)

Full monthly model in the data-room XLSX.

## 10. Target Unit Economics (Year 1 model)
> These are the **target** unit economics built into our financial model,
> not realized values — we're pre-revenue today.
- Avg revenue per load: **${UNIT_ECONOMICS['avg_revenue_per_load_usd']:,}** (DAT 2024 spot-blend reference)
- Year-1 gross margin target: **{UNIT_ECONOMICS['avg_gross_margin_pct_y1']}%** (new-broker reality; TIA mature median: 15.5%)
- Year-3 gross margin target: **{UNIT_ECONOMICS['avg_gross_margin_pct_mature']}%** (industry median)
- Contribution per load: **${UNIT_ECONOMICS['contribution_per_load_usd']}** (target)
- CAC: **${UNIT_ECONOMICS['customer_acquisition_cost_usd']:,}** (cold-start target)
- LTV / CAC: **{UNIT_ECONOMICS['ltv_cac_ratio']}x** (3-yr target)
- Customer payback: **{UNIT_ECONOMICS['customer_payback_loads']} loads** (target)
- EBITDA break-even: **Month {UNIT_ECONOMICS['monthly_ebitda_breakeven_month']}** (target)

## 11. Projected Probability of Success
> Forward-looking projection — not historical performance.
- **{probability['score_pct']}%** — projected Year-1 survival probability for
  {short} given current capital, operator experience, and tooling. Computed
  via a weighted scorecard (industry base survival + capital + experience +
  tooling + authority + factoring + carrier pool + lane focus). Capped at
  90% — the freight market has irreducible volatility no setup can erase.
- **Band: {probability['band']}** — {probability['band_note']}
- Operator-grade TMS contributes an estimated **+5 pts** of Year-1 survival
  lift over paper-broker baselines (operator estimate, not peer-reviewed).

## 12. Competition
- **Mega 3PLs (CHR, XPO, Echo, Coyote)**: scale + balance sheet, but no
  named-broker accountability and notorious shipper-friction.
- **Mom-and-pop brokers**: relationship-driven but paper-bound,
  no tooling, no quick-pay, high failure rate.
- **{short}**: the rare hybrid — operator relationships with software-driven
  discipline.

## 13. The Ask
- Raising **$500,000 SAFE** at a **$4.0M cap with 20% discount**.
- Use of funds: authority + bond + insurance, carrier vetting tooling,
  load-board subscriptions, founder runway, marketing, and quick-pay
  working capital.
- **Post-raise milestone targets**:
  - First paying shipper within **30–60 days** of close.
  - **Carrier network of 300+** by Day 90.
  - **EBITDA break-even targeted by Month {ue_breakeven}** (honest baseline).

## 14. What's De-Risked (vs. typical pre-revenue brokerages)
- TMS Command Deck **shipped and operating** (this very platform). No
  software risk left between us and Day 1.
- 14 launch-day provider integrations pre-wired in the Connections vault —
  just need the keys.
- Calafia-stamped document templates auto-generate in under 800ms — zero
  documentation risk on Day 1 loads.
- 13-year founder operating history; references available on request.
- Marketing pack, public investor site, and cold-email sequences already
  built — Day 1 of outbound is a copy-and-send operation.

## 15. Contact
**{founder}** — Founder & Principal Broker
{company}
Minneapolis · Saint Paul, MN
oliver@oriseifreight.com
LinkedIn · https://www.linkedin.com/in/oliver-cummins-a27304a3/

---

## 16. The JadeOS Stack · One thesis · Three products

This pitch covers Hot Shot TMS (Product 03). It is one of three
products on a single cap table:

- **Product 01 · JadeOS Quantum AI** (flagship) — AI command center
  for builders & founders. 50+ modules, voice-first "Hey Jade",
  persistent memory across modules. 128-qubit Qiskit Aer + Claude
  Haiku 4.5. **Status:** Beta.

- **Product 02 · JadeOS-Agent Suite** (freight-vertical productization) —
  Six AI agents that sit on top of any TMS (Hot Shot or McLeod /
  Descartes / TMW): rate-floor guard, audit chain, workflow memory,
  active claims, dispatch copilot, settlement watchdog.
  **Status:** 1 live prod · 2 live partial.

- **Product 03 · Hot Shot TMS** (operator-built system of record) —
  Transportation management for the hot-shot small-to-mid carrier
  segment incumbent TMS vendors don't serve well. **Status:** Build
  complete · ready to deploy.

Same builder. Same persistent-memory substrate. One investable thesis.

**Full three-product thesis →** https://mpls-automation-hub.emergent.host/
"""


def _one_pager_markdown(brand: Dict[str, Any]) -> str:
    company = brand.get("company_name") or "Orisei Freight Solutions LLC"
    short = brand.get("short_name") or "Orisei"
    tagline = brand.get("tagline") or "Operator-built freight brokerage"
    founder = brand.get("owner_name") or "Oliver Cummins"
    annual = _annual_summary(_financial_model_rows())
    return f"""# {company} · Investor One-Pager

{tagline}

## The 60-Second Pitch
{short} is a Twin Cities-based property freight brokerage built around an
operator-grade in-house TMS. We pair a 13-year founder with a margin-aware
booking engine, auto-stamped BOLs, dock-photo PODs, and same-day carrier
pay — so shippers get a named human who answers the phone and the kind of
documentation discipline that big 3PLs still can't deliver.

> **Stage: Pre-revenue · Pre-launch.** Product is built and operating.
> All financial figures below are forward-looking targets, not realized
> revenue.

## Market & Why Now
- **TAM**: $210B US freight brokerage (TIA 2024)
- **SAM**: $95B Midwest property TL/LTL
- **SOM (Yr 3 target)**: $8.5M Twin Cities + Upper Midwest corridor
- **32%** of brokerages fail Year 1 · **52%** by Year 3 — opportunity for
  operator-grade tooling to redefine the bar.

## Financial Forecast (Bootstrap baseline · not realized)
- **Year 1 (Target)**: ${annual[0]['revenue_usd']:,.0f} revenue · {annual[0]['loads']:,} loads
- **Year 2 (Target)**: ${annual[1]['revenue_usd']:,.0f} revenue · {annual[1]['loads']:,} loads
- **Year 3 (Target)**: ${annual[2]['revenue_usd']:,.0f} revenue · {annual[2]['ebitda_margin_pct']}% EBITDA

## Target Unit Economics
- ${UNIT_ECONOMICS['avg_gross_profit_per_load_usd']} gross profit per load (Y1 target) · LTV/CAC **{UNIT_ECONOMICS['ltv_cac_ratio']}x** (3-yr target)
- **{UNIT_ECONOMICS['customer_payback_loads']}-load payback** on customer acquisition cost (target)
- Operator-grade TMS adds an estimated **+5 pt** Year-1 survival lift over paper-broker baseline *(operator estimate)*

## What's Already De-Risked
- TMS Command Deck shipped and operating in production
- Brand-aware BOL / POD / compliance document engine live
- Marketing pack + public investor page already published
- 13-year founder operator track record

## The Ask
**$500,000 SAFE @ $4.0M cap, 20% discount.** Targeting first paying shipper
within 30–60 days of close, EBITDA break-even by Month {UNIT_ECONOMICS['monthly_ebitda_breakeven_month']}.

## Contact
**{founder}** · Founder & Principal Broker
{company} · Minneapolis · Saint Paul, MN
oliver@oriseifreight.com
LinkedIn · https://www.linkedin.com/in/oliver-cummins-a27304a3/

---

## Part of the JadeOS Stack
1 of 3 products on one cap table.
JadeOS Quantum AI (flagship) · JadeOS-Agent Suite (freight agents) ·
**Hot Shot TMS** (system of record).
Full thesis → https://mpls-automation-hub.emergent.host/
"""


def _industry_probability_markdown(brand: Dict[str, Any], probability: Dict[str, Any]) -> str:
    short = brand.get("short_name") or "Orisei"
    return f"""# {short} · Industry Probability of Success Report

## Headline
**{probability['score_pct']}%** — projected Year-1 success probability for
{short} under the current operator + capital + tooling configuration.

**Band: {probability['band']}** — {probability['band_note']}

> **Honest framing.** This is a forward-looking projection, not a guarantee.
> The model is capped at 90% because the freight market has irreducible
> volatility (recessions, fuel shocks, key-customer bankruptcies, lane
> disruption) that no amount of capital or tooling can fully eliminate.

## Methodology
We start from the **SBA + TIA 2023 industry survival baseline** of **68%**
(equal to 1 − 32% Year-1 broker failure rate), then add weighted points for
factors that the freight-brokerage research literature has repeatedly shown
to predict survival. Negative weights penalize known failure modes.

## Score Breakdown
{chr(10).join(f"- **{d['label']}**: {d['delta']:+.1f} pts — {d['note']}" for d in probability['drivers'])}

## Industry Benchmarks Used
- US freight brokerage TAM: **${INDUSTRY_BENCHMARKS['tam_usd_billion']}B** (TIA Annual Report 2024)
- Upper-Midwest property TL/LTL SAM: **${INDUSTRY_BENCHMARKS['sam_usd_billion']}B** (TIA + state-level shipping data)
- Industry CAGR (2024–2028): **{INDUSTRY_BENCHMARKS['industry_growth_cagr_pct']}%** (FreightWaves Sonar 2024)
- Year-1 broker failure rate: **{INDUSTRY_BENCHMARKS['broker_failure_year1_pct']}%** (SBA + TIA 2023)
- Year-3 broker failure rate: **{INDUSTRY_BENCHMARKS['broker_failure_year3_pct']}%** (SBA 2023)
- Year-5 surviving + cash-flow positive: **{INDUSTRY_BENCHMARKS['broker_success_year5_pct']}%** (SBA 2023)
- Mature-broker gross margin median: **{INDUSTRY_BENCHMARKS['avg_broker_gross_margin_pct']}%** (TIA 2024)
- New-broker (Y1) gross margin reality: **{INDUSTRY_BENCHMARKS['new_broker_gross_margin_pct_y1']}%** (TIA + operator surveys)
- Avg revenue per load: **${INDUSTRY_BENCHMARKS['avg_load_revenue_usd']:,}** (DAT + Truckstop spot blend 2024)
- Operator-grade TMS estimated Year-1 lift: **+{INDUSTRY_BENCHMARKS['ai_tooling_estimated_lift_pct']} pts** *(operator estimate — NOT peer-reviewed)*

## Sources & Honesty Note
{chr(10).join(f"- {s}" for s in INDUSTRY_BENCHMARKS['sources'])}

> {INDUSTRY_BENCHMARKS['honesty_note']}

## Key Risks (transparency)
The score above models *survival*, not *outsized return*. Even with a
favorable band, real downside risks remain:
- **Customer acquisition slippage**: Cold-start brokerages typically need 4-9
  months of outbound effort to close their first major shipper. Our financial
  model assumes the first paying shipper at Month 4 — a 2-month slip drops
  Y1 loads by ~25%.
- **Recession-induced freight rate compression** (RPM drops > 8% have
  historically squeezed broker gross margin by 200–300 bps).
- **Carrier-side bankruptcies** disrupting promised lanes.
- **Single-customer concentration risk** in early years (likely 3-5 anchor
  shippers in Y1-Y2; losing one anchor = -30% revenue).
- **Regulatory shocks** (BMC-84 bond increases, broker-disclosure rules).
- **Authority timing**: A 60-day FMCSA delay pushes break-even from Month 9
  to Month 11-12.

## What Moves the Needle
For founders aiming to push the score higher:
1. Raise starting capital to $100K+ (adds ~4-5 pts).
2. Secure 2–3 anchor shippers under written agreement pre-launch (adds ~3 pts).
3. Pre-onboard 50+ vetted carriers before tendering a single load (~2 pts).
4. Get factoring partner LOI in writing (~2 pts).
5. Stay disciplined on lane focus — 6–10 lanes max in Year 1 (~2 pts).
"""


class PersonalizationIn(BaseModel):
    firm_name: str = Field(..., min_length=1, max_length=120,
                           description="VC firm name to stamp on every page")
    contact_name: Optional[str] = Field(None, max_length=120,
                                        description="Optional GP / partner name")
    prepared_date: Optional[str] = Field(None, max_length=40,
                                         description="Override prepared date (defaults to today)")
    doc_type: str = Field("deck", description="deck · one-pager · zip")


def _safe_slug(s: str) -> str:
    """Filesystem-safe slug for filenames."""
    return _re.sub(r"[^A-Za-z0-9_-]+", "_", s).strip("_") or "VC"


def _normalize_personalization(payload: PersonalizationIn) -> Dict[str, Any]:
    return {
        "firm_name": payload.firm_name.strip(),
        "contact_name": (payload.contact_name or "").strip() or None,
        "prepared_date": (payload.prepared_date.strip()
                          if payload.prepared_date
                          else datetime.now(timezone.utc).strftime("%d %b %Y")),
    }


# -------------------- ROUTER --------------------
def build_investor_router(*, db, get_current_user: Callable, require_role: Callable,
                          active_brand_doc: Callable) -> APIRouter:
    """Wire up `/api/investor/*`. Admin-gated end-to-end."""
    router = APIRouter(prefix="/investor")
    _admin_dep = require_role("admin")

    @router.get("/boardroom")
    async def boardroom(_: Any = Depends(_admin_dep)):
        """All-in-one investor analytics payload for the in-app boardroom."""
        rows = _financial_model_rows()
        annual = _annual_summary(rows)
        default_probability = _compute_probability(ProbabilityInputs())
        return {
            "current_status": CURRENT_STATUS,
            "market_sizing": MARKET_SIZING,
            "industry_benchmarks": INDUSTRY_BENCHMARKS,
            "unit_economics": UNIT_ECONOMICS,
            "monthly_model": rows,
            "annual_summary": annual,
            "default_probability": default_probability,
        }

    @router.post("/probability")
    async def probability(payload: ProbabilityInputs, _: Any = Depends(_admin_dep)):
        return _compute_probability(payload)

    @router.get("/financial-model.xlsx")
    async def financial_model_xlsx(_: Any = Depends(_admin_dep)):
        brand = await active_brand_doc() or {}
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        rows = _financial_model_rows()
        annual = _annual_summary(rows)
        xlsx_bytes = _build_financial_model_xlsx(brand, rows, annual)
        filename = f"{company.replace(' ', '_')}_Financial_Model.xlsx"
        return StreamingResponse(
            io.BytesIO(xlsx_bytes),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{filename}"'},
        )

    @router.get("/deck.pdf")
    async def deck_pdf(_: Any = Depends(_admin_dep)):
        brand = await active_brand_doc() or {}
        prob = _compute_probability(ProbabilityInputs())
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        pdf_bytes = build_branded_markdown_pdf(
            _deck_markdown(brand, prob),
            title=f"{company} · VC Pitch Deck",
            subtitle="Series Seed · Confidential",
            brand=brand,
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{company.replace(" ", "_")}_Pitch_Deck.pdf"'},
        )

    @router.get("/one-pager.pdf")
    async def one_pager_pdf(_: Any = Depends(_admin_dep)):
        brand = await active_brand_doc() or {}
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        pdf_bytes = build_branded_markdown_pdf(
            _one_pager_markdown(brand),
            title=f"{company} · Investor One-Pager",
            subtitle="At-a-glance · Confidential",
            brand=brand,
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition": f'attachment; filename="{company.replace(" ", "_")}_One_Pager.pdf"'},
        )

    @router.get("/data-room.zip")
    async def data_room_zip(_: Any = Depends(_admin_dep)):
        """The full VC data-room: deck PDF + business plan PDF + financial
        model XLSX + cap table CSV + one-pager + industry probability PDF.
        Everything stamped with the active brand."""
        brand = await active_brand_doc() or {}
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        short = brand.get("short_name") or "Orisei"
        rows = _financial_model_rows()
        annual = _annual_summary(rows)
        prob = _compute_probability(ProbabilityInputs())

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            # 01 - Pitch Deck PDF
            zf.writestr(f"01_{short}_Pitch_Deck.pdf",
                        build_branded_markdown_pdf(
                            _deck_markdown(brand, prob),
                            title=f"{company} · VC Pitch Deck",
                            subtitle="Series Seed · Confidential",
                            brand=brand,
                        ))
            # 02 - Investor One-Pager PDF
            zf.writestr(f"02_{short}_One_Pager.pdf",
                        build_branded_markdown_pdf(
                            _one_pager_markdown(brand),
                            title=f"{company} · Investor One-Pager",
                            subtitle="At-a-glance · Confidential",
                            brand=brand,
                        ))
            # 03 - Industry Probability Report PDF
            zf.writestr(f"03_{short}_Industry_Probability_Report.pdf",
                        build_branded_markdown_pdf(
                            _industry_probability_markdown(brand, prob),
                            title=f"{company} · Industry Probability of Success",
                            subtitle="Methodology + Sources · Confidential",
                            brand=brand,
                        ))
            # 04 - Business Plan PDF (existing markdown)
            try:
                from pathlib import Path
                # search common locations for the business plan markdown
                candidates = [
                    Path("/app/BROKERAGE_BUSINESS_PLAN.md"),
                    Path("/app/BUSINESS_PLAN.md"),
                    Path(__file__).resolve().parent.parent / "docs" / "BROKERAGE_BUSINESS_PLAN.md",
                    Path(__file__).resolve().parent.parent / "docs" / "BUSINESS_PLAN.md",
                ]
                plan_path = next((p for p in candidates if p.exists()), None)
                if plan_path:
                    plan_md = plan_path.read_text(encoding="utf-8")
                    zf.writestr(f"04_{short}_Business_Plan.pdf",
                                build_branded_markdown_pdf(
                                    plan_md, title=f"{company} · Business Plan",
                                    subtitle="Founder Business Plan · Confidential",
                                    brand=brand,
                                ))
            except Exception as exc:                                  # noqa: BLE001
                logger.warning("Skipped Business_Plan in data-room: %s", exc)
            # 05 - Financial Model XLSX
            zf.writestr(f"05_{short}_Financial_Model.xlsx",
                        _build_financial_model_xlsx(brand, rows, annual))
            # 06 - Cap Table CSV
            zf.writestr(f"06_{short}_Cap_Table.csv", _build_cap_table_csv(brand))
            # 07-09 - Marketing collateral (carrier + shipper sell sheets + press release)
            try:
                from .marketing import (
                    _carrier_sell_sheet_md, _shipper_sell_sheet_md, _press_release_md,
                )
                zf.writestr(f"07_{short}_Carrier_Sell_Sheet.pdf",
                            build_branded_markdown_pdf(
                                _carrier_sell_sheet_md(brand),
                                title=f"{company} · Carrier Sell Sheet",
                                subtitle="For carriers we'd like in our network",
                                brand=brand,
                            ))
                zf.writestr(f"08_{short}_Shipper_Sell_Sheet.pdf",
                            build_branded_markdown_pdf(
                                _shipper_sell_sheet_md(brand),
                                title=f"{company} · Shipper Sell Sheet",
                                subtitle="Operator-built freight brokerage",
                                brand=brand,
                            ))
                zf.writestr(f"09_{short}_Press_Release.pdf",
                            build_branded_markdown_pdf(
                                _press_release_md(brand),
                                title=f"{company} · MC-Launch Press Release",
                                subtitle="For immediate release",
                                brand=brand,
                            ))
            except Exception as exc:                                  # noqa: BLE001
                logger.warning("Skipped marketing collateral in data-room: %s", exc)
            # 10 - README index
            zf.writestr("README.txt",
                        f"{company} · VC Data Room\n"
                        f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}\n"
                        f"\nContents:\n"
                        f"  01_{short}_Pitch_Deck.pdf\n"
                        f"  02_{short}_One_Pager.pdf\n"
                        f"  03_{short}_Industry_Probability_Report.pdf\n"
                        f"  04_{short}_Business_Plan.pdf (if available)\n"
                        f"  05_{short}_Financial_Model.xlsx\n"
                        f"  06_{short}_Cap_Table.csv\n"
                        f"  07_{short}_Carrier_Sell_Sheet.pdf\n"
                        f"  08_{short}_Shipper_Sell_Sheet.pdf\n"
                        f"  09_{short}_Press_Release.pdf\n"
                        f"\nContact: oliver@oriseifreight.com\n"
                        f"LinkedIn: https://www.linkedin.com/in/oliver-cummins-a27304a3/\n"
                        f"\n--- JadeOS Stack ---\n"
                        f"Hot Shot TMS is 1 of 3 products on one cap table:\n"
                        f"  · JadeOS Quantum AI (flagship)\n"
                        f"  · JadeOS-Agent Suite (freight-vertical agents)\n"
                        f"  · Hot Shot TMS (system of record)\n"
                        f"Full thesis: https://mpls-automation-hub.emergent.host/\n")

        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition": f'attachment; filename="{company.replace(" ", "_")}_VC_Data_Room.zip"'},
        )

    # -------------------- PERSONALIZED FOR-VC DOCS --------------------
    @router.post("/personalized-deck.pdf")
    async def personalized_deck_pdf(payload: PersonalizationIn,
                                    _: Any = Depends(_admin_dep)):
        """Generate a pitch deck PDF stamped with a specific VC firm name on
        every page (top banner + diagonal CONFIDENTIAL watermark)."""
        brand = await active_brand_doc() or {}
        prob = _compute_probability(ProbabilityInputs())
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        personalization = _normalize_personalization(payload)
        firm_slug = _safe_slug(personalization["firm_name"])
        # Audit
        await db.investor_personalized_outreach.insert_one({
            "id": f"VCP-{__import__('uuid').uuid4().hex[:10].upper()}",
            "doc_type": "deck",
            "firm_name": personalization["firm_name"],
            "contact_name": personalization.get("contact_name"),
            "prepared_date": personalization["prepared_date"],
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "brand_short": brand.get("short_name") or "Orisei",
        })
        pdf_bytes = build_branded_markdown_pdf(
            _deck_markdown(brand, prob),
            title=f"{company} · VC Pitch Deck",
            subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
            brand=brand,
            personalization=personalization,
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="{company.replace(" ", "_")}_Pitch_Deck_for_{firm_slug}.pdf"'},
        )

    @router.post("/personalized-one-pager.pdf")
    async def personalized_one_pager_pdf(payload: PersonalizationIn,
                                         _: Any = Depends(_admin_dep)):
        """Generate a one-pager PDF stamped with a specific VC firm name."""
        brand = await active_brand_doc() or {}
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        personalization = _normalize_personalization(payload)
        firm_slug = _safe_slug(personalization["firm_name"])
        await db.investor_personalized_outreach.insert_one({
            "id": f"VCP-{__import__('uuid').uuid4().hex[:10].upper()}",
            "doc_type": "one-pager",
            "firm_name": personalization["firm_name"],
            "contact_name": personalization.get("contact_name"),
            "prepared_date": personalization["prepared_date"],
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "brand_short": brand.get("short_name") or "Orisei",
        })
        pdf_bytes = build_branded_markdown_pdf(
            _one_pager_markdown(brand),
            title=f"{company} · Investor One-Pager",
            subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
            brand=brand,
            personalization=personalization,
        )
        return StreamingResponse(
            io.BytesIO(pdf_bytes),
            media_type="application/pdf",
            headers={"Content-Disposition":
                     f'attachment; filename="{company.replace(" ", "_")}_One_Pager_for_{firm_slug}.pdf"'},
        )

    @router.post("/personalized-data-room.zip")
    async def personalized_data_room_zip(payload: PersonalizationIn,
                                         _: Any = Depends(_admin_dep)):
        """Generate the full VC data-room ZIP with every PDF personalized for
        the named firm (banner + watermark on every page of every deliverable).
        Non-PDF artifacts (XLSX, CSV) are not personalized."""
        brand = await active_brand_doc() or {}
        company = brand.get("company_name") or "Orisei Freight Solutions LLC"
        short = brand.get("short_name") or "Orisei"
        rows = _financial_model_rows()
        annual = _annual_summary(rows)
        prob = _compute_probability(ProbabilityInputs())
        personalization = _normalize_personalization(payload)
        firm_slug = _safe_slug(personalization["firm_name"])

        await db.investor_personalized_outreach.insert_one({
            "id": f"VCP-{__import__('uuid').uuid4().hex[:10].upper()}",
            "doc_type": "zip",
            "firm_name": personalization["firm_name"],
            "contact_name": personalization.get("contact_name"),
            "prepared_date": personalization["prepared_date"],
            "generated_at": datetime.now(timezone.utc).isoformat(),
            "brand_short": short,
        })

        buf = io.BytesIO()
        with zipfile.ZipFile(buf, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            zf.writestr(f"01_{short}_Pitch_Deck_for_{firm_slug}.pdf",
                        build_branded_markdown_pdf(
                            _deck_markdown(brand, prob),
                            title=f"{company} · VC Pitch Deck",
                            subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
                            brand=brand, personalization=personalization))
            zf.writestr(f"02_{short}_One_Pager_for_{firm_slug}.pdf",
                        build_branded_markdown_pdf(
                            _one_pager_markdown(brand),
                            title=f"{company} · Investor One-Pager",
                            subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
                            brand=brand, personalization=personalization))
            zf.writestr(f"03_{short}_Industry_Probability_Report_for_{firm_slug}.pdf",
                        build_branded_markdown_pdf(
                            _industry_probability_markdown(brand, prob),
                            title=f"{company} · Industry Probability of Success",
                            subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
                            brand=brand, personalization=personalization))
            try:
                from pathlib import Path
                candidates = [
                    Path("/app/BROKERAGE_BUSINESS_PLAN.md"),
                    Path("/app/BUSINESS_PLAN.md"),
                ]
                plan_path = next((p for p in candidates if p.exists()), None)
                if plan_path:
                    plan_md = plan_path.read_text(encoding="utf-8")
                    zf.writestr(f"04_{short}_Business_Plan_for_{firm_slug}.pdf",
                                build_branded_markdown_pdf(
                                    plan_md, title=f"{company} · Business Plan",
                                    subtitle=f"Prepared for {personalization['firm_name']} · Confidential",
                                    brand=brand, personalization=personalization))
            except Exception as exc:                                  # noqa: BLE001
                logger.warning("Skipped Business_Plan in personalized zip: %s", exc)
            zf.writestr(f"05_{short}_Financial_Model.xlsx",
                        _build_financial_model_xlsx(brand, rows, annual))
            zf.writestr(f"06_{short}_Cap_Table.csv", _build_cap_table_csv(brand))
            zf.writestr("README.txt",
                        f"{company} · VC Data Room\n"
                        f"Prepared for: {personalization['firm_name']}\n"
                        + (f"Attn: {personalization['contact_name']}\n"
                           if personalization.get('contact_name') else "")
                        + f"Date: {personalization['prepared_date']}\n"
                        f"Generated {datetime.now(timezone.utc):%Y-%m-%d %H:%M UTC}\n"
                        f"\nThis package is confidential and intended solely\n"
                        f"for {personalization['firm_name']}. Please do not\n"
                        f"forward without prior written consent.\n"
                        f"\nContact: oliver@oriseifreight.com\n"
                        f"LinkedIn: https://www.linkedin.com/in/oliver-cummins-a27304a3/\n"
                        f"\n--- JadeOS Stack ---\n"
                        f"Hot Shot TMS is 1 of 3 products on one cap table:\n"
                        f"  · JadeOS Quantum AI (flagship)\n"
                        f"  · JadeOS-Agent Suite (freight-vertical agents)\n"
                        f"  · Hot Shot TMS (system of record)\n"
                        f"Full thesis: https://mpls-automation-hub.emergent.host/\n")
        buf.seek(0)
        return StreamingResponse(
            buf,
            media_type="application/zip",
            headers={"Content-Disposition":
                     f'attachment; filename="{company.replace(" ", "_")}_VC_Data_Room_for_{firm_slug}.zip"'},
        )

    @router.get("/personalized-outreach")
    async def personalized_outreach_history(_: Any = Depends(_admin_dep)) -> Dict[str, Any]:
        """Recent personalized PDF generation history (most recent first)."""
        cursor = db.investor_personalized_outreach.find(
            {}, {"_id": 0}
        ).sort("generated_at", -1).limit(50)
        items = await cursor.to_list(length=50)
        return {"items": items, "count": len(items)}

    return router

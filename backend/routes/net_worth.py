"""routes.net_worth — Orisei partnership net worth statements (insurance underwriting)."""
from __future__ import annotations

import io
import uuid
from datetime import datetime, timezone
from typing import Any, Callable, Dict, List, Optional

from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import Response
from pydantic import BaseModel, Field
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen.canvas import Canvas

W, H = letter
PAPER = colors.HexColor("#FAF6ED")
INK = colors.HexColor("#1C2430")
AZURE = colors.HexColor("#123B5C")
GOLD = colors.HexColor("#C9A227")
SLATE = colors.HexColor("#5B6472")
LINE = colors.HexColor("#D8CFBA")

ASSET_CATEGORIES = ["Cash & bank accounts", "Investments & retirement (401k/IRA/brokerage)",
                    "Real estate (market value)", "Vehicles & equipment",
                    "Business interests / ownership stakes", "Other assets"]
LIABILITY_CATEGORIES = ["Mortgages / real estate loans", "Vehicle & equipment loans",
                        "Credit cards & revolving debt", "Student / personal loans",
                        "Other liabilities"]
MEMBERS = ["Oliver Cummins", "Daniel W. Karsor", "Doug Graham"]


def _now() -> str:
    return datetime.now(timezone.utc).isoformat()


class LineItem(BaseModel):
    category: str
    description: str = ""
    value: float = Field(0, ge=0)


class MemberStatement(BaseModel):
    member_name: str = Field(..., min_length=1, max_length=80)
    as_of_date: str = ""
    assets: List[LineItem] = []
    liabilities: List[LineItem] = []
    status: str = "draft"  # draft | submitted


def _totals(doc: Dict[str, Any]) -> Dict[str, float]:
    a = round(sum(x.get("value", 0) for x in doc.get("assets", [])), 2)
    li = round(sum(x.get("value", 0) for x in doc.get("liabilities", [])), 2)
    return {"total_assets": a, "total_liabilities": li, "net_worth": round(a - li, 2)}


def _header(c: Canvas, title: str, subtitle: str):
    c.setFillColor(PAPER)
    c.rect(0, 0, W, H, fill=1, stroke=0)
    c.setFillColor(AZURE)
    c.rect(0, H - 96, W, 96, fill=1, stroke=0)
    c.setFillColor(GOLD)
    c.rect(0, H - 100, W, 4, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 19)
    c.setFillColor(colors.white)
    c.drawString(46, H - 46, "ORISEI FREIGHT SOLUTIONS LLC")
    c.setFont("Helvetica-Bold", 12)
    c.setFillColor(GOLD)
    c.drawString(46, H - 66, title)
    c.setFont("Helvetica", 8.5)
    c.setFillColor(colors.HexColor("#B9C6D4"))
    c.drawString(46, H - 82, subtitle)
    c.setFont("Helvetica", 8)
    c.drawRightString(W - 46, H - 46, "Minneapolis–St. Paul, MN")
    c.drawRightString(W - 46, H - 58, "(763) 443-4459")
    c.drawRightString(W - 46, H - 70, "oliver@oriseifreightsolutions.com")


def _footer(c: Canvas, note: str):
    c.setFont("Helvetica", 7.5)
    c.setFillColor(SLATE)
    c.drawCentredString(W / 2, 30, note)
    c.setFillColor(GOLD)
    c.rect(0, 20, W, 1.5, fill=1, stroke=0)


def _rule_rows(c: Canvas, x, y, w, categories, blank_rows_per=2):
    for cat in categories:
        c.setFont("Helvetica-Bold", 9.5)
        c.setFillColor(AZURE)
        c.drawString(x, y, cat.upper())
        c.setFont("Helvetica", 8)
        c.setFillColor(SLATE)
        c.drawRightString(x + w, y, "VALUE ($)")
        y -= 6
        for _ in range(blank_rows_per):
            y -= 18
            c.setStrokeColor(LINE)
            c.setLineWidth(0.7)
            c.line(x, y, x + w * 0.68, y)
            c.line(x + w * 0.74, y, x + w, y)
        y -= 16
    return y


def build_template_pdf() -> bytes:
    buf = io.BytesIO()
    c = Canvas(buf, pagesize=letter)
    c.setTitle("Orisei · Member Net Worth Statement Template")
    # Page 1 — instructions + identity + assets
    _header(c, "MEMBER PERSONAL NET WORTH STATEMENT",
            "Confidential — for partnership insurance underwriting (BMC-84 surety, liability & cargo programs)")
    y = H - 126
    c.setFont("Helvetica", 9)
    c.setFillColor(INK)
    for ln in ["Instructions: list current market values as of the statement date. Estimate honestly — underwriters",
               "verify large items. Attach recent statements where available. Return the completed form to Oliver",
               "Cummins or enter it directly in the Command Deck under Brokerage → Net Worth."]:
        c.drawString(46, y, ln)
        y -= 13
    y -= 8
    c.setFont("Helvetica-Bold", 9.5)
    c.setFillColor(AZURE)
    for label in ["MEMBER NAME", "STATEMENT DATE (AS OF)", "PHONE / EMAIL"]:
        c.drawString(46, y, label)
        c.setStrokeColor(LINE)
        c.line(196, y - 2, W - 46, y - 2)
        y -= 24
    y -= 6
    c.setFillColor(GOLD)
    c.rect(46, y - 4, W - 92, 20, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(AZURE)
    c.drawString(56, y + 1, "SECTION 1 · ASSETS (what you own)")
    y -= 26
    y = _rule_rows(c, 46, y, W - 92, ASSET_CATEGORIES, 2)
    c.setFont("Helvetica-Bold", 10)
    c.setFillColor(INK)
    c.drawString(46, y, "TOTAL ASSETS")
    c.setStrokeColor(INK)
    c.line(W - 180, y - 2, W - 46, y - 2)
    _footer(c, "Orisei Freight Solutions LLC · Member-confidential · Page 1 of 2")
    c.showPage()
    # Page 2 — liabilities + certification
    _header(c, "MEMBER PERSONAL NET WORTH STATEMENT", "Section 2 — Liabilities & certification")
    y = H - 130
    c.setFillColor(GOLD)
    c.rect(46, y - 4, W - 92, 20, fill=1, stroke=0)
    c.setFont("Helvetica-Bold", 11)
    c.setFillColor(AZURE)
    c.drawString(56, y + 1, "SECTION 2 · LIABILITIES (what you owe)")
    y -= 26
    y = _rule_rows(c, 46, y, W - 92, LIABILITY_CATEGORIES, 2)
    for label in ["TOTAL LIABILITIES", "NET WORTH (assets − liabilities)"]:
        c.setFont("Helvetica-Bold", 10)
        c.setFillColor(INK)
        c.drawString(46, y, label)
        c.setStrokeColor(INK)
        c.line(W - 180, y - 2, W - 46, y - 2)
        y -= 26
    y -= 10
    c.setFont("Helvetica", 8.5)
    c.setFillColor(INK)
    for ln in ["I certify that the information above is true and complete to the best of my knowledge, provided for",
               "the purpose of insurance and surety underwriting for Orisei Freight Solutions LLC."]:
        c.drawString(46, y, ln)
        y -= 12
    y -= 22
    for label in ["MEMBER SIGNATURE", "DATE"]:
        c.setStrokeColor(INK)
        c.line(46 if label == "MEMBER SIGNATURE" else 340, y, (300 if label == "MEMBER SIGNATURE" else W - 46), y)
        c.setFont("Helvetica", 7.5)
        c.setFillColor(SLATE)
        c.drawString(46 if label == "MEMBER SIGNATURE" else 340, y - 11, label)
    _footer(c, "Orisei Freight Solutions LLC · Member-confidential · Page 2 of 2")
    c.save()
    return buf.getvalue()


def build_master_pdf(members: List[Dict[str, Any]]) -> bytes:
    buf = io.BytesIO()
    c = Canvas(buf, pagesize=letter)
    c.setTitle("Orisei · Partnership Net Worth Statement (Master)")
    _header(c, "PARTNERSHIP NET WORTH STATEMENT — MASTER",
            f"Combined member statements · prepared {datetime.now(timezone.utc).strftime('%B %d, %Y')} · for insurance & surety underwriting")
    y = H - 130
    combined_a = sum(_totals(m)["total_assets"] for m in members)
    combined_l = sum(_totals(m)["total_liabilities"] for m in members)
    # combined banner
    c.setFillColor(AZURE)
    c.roundRect(46, y - 58, W - 92, 58, 8, fill=1, stroke=0)
    for i, (lab, val) in enumerate([("COMBINED ASSETS", combined_a), ("COMBINED LIABILITIES", combined_l),
                                    ("COMBINED NET WORTH", combined_a - combined_l)]):
        cx = 46 + (W - 92) / 6 + i * (W - 92) / 3
        c.setFont("Helvetica-Bold", 15)
        c.setFillColor(GOLD if i == 2 else colors.white)
        c.drawCentredString(cx, y - 30, f"${val:,.0f}")
        c.setFont("Helvetica", 7)
        c.setFillColor(colors.HexColor("#B9C6D4"))
        c.drawCentredString(cx, y - 46, lab)
    y -= 84
    for m in members:
        t = _totals(m)
        need = 66 + 13 * (len(m.get("assets", [])) + len(m.get("liabilities", [])))
        if y - need < 70:
            _footer(c, "Orisei Freight Solutions LLC · Confidential underwriting document")
            c.showPage()
            _header(c, "PARTNERSHIP NET WORTH STATEMENT — MASTER", "Continued")
            y = H - 130
        c.setFillColor(GOLD)
        c.rect(46, y - 2, W - 92, 20, fill=1, stroke=0)
        c.setFont("Helvetica-Bold", 10.5)
        c.setFillColor(AZURE)
        c.drawString(54, y + 3, f"{m['member_name'].upper()} — 33⅓% MEMBER")
        c.setFont("Helvetica", 8)
        c.drawRightString(W - 54, y + 3, f"as of {m.get('as_of_date') or '—'} · {m.get('status', 'draft').upper()}")
        y -= 20
        for section, items, col in [("ASSETS", m.get("assets", []), colors.HexColor("#1F6B3A")),
                                    ("LIABILITIES", m.get("liabilities", []), colors.HexColor("#8A2E1F"))]:
            c.setFont("Helvetica-Bold", 8)
            c.setFillColor(col)
            c.drawString(54, y - 12, section)
            y -= 14
            if not items:
                c.setFont("Helvetica-Oblique", 8)
                c.setFillColor(SLATE)
                c.drawString(70, y - 11, "No entries yet")
                y -= 13
            for it in items:
                c.setFont("Helvetica", 8.5)
                c.setFillColor(INK)
                desc = f"{it['category']}" + (f" — {it['description']}" if it.get("description") else "")
                c.drawString(70, y - 11, desc[:88])
                c.drawRightString(W - 54, y - 11, f"${it.get('value', 0):,.0f}")
                y -= 13
        c.setFont("Helvetica-Bold", 9)
        c.setFillColor(INK)
        c.drawString(54, y - 13, "MEMBER NET WORTH")
        c.drawRightString(W - 54, y - 13, f"${t['net_worth']:,.0f}")
        c.setStrokeColor(LINE)
        c.line(46, y - 20, W - 46, y - 20)
        y -= 34
    c.setFont("Helvetica", 7.5)
    c.setFillColor(SLATE)
    c.drawString(46, max(y - 6, 46), "Each member statement is certified individually on the Orisei Member Net Worth Statement form. Values are member-reported market values.")
    _footer(c, "Orisei Freight Solutions LLC · Confidential underwriting document")
    c.save()
    return buf.getvalue()


def build_template_xlsx(member_name: str = "") -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    AZ, GD, PP, LN = "123B5C", "C9A227", "FAF6ED", "D8CFBA"
    wb = Workbook()
    ws = wb.active
    ws.title = "Net Worth Statement"
    ws.sheet_view.showGridLines = False
    for col, w in zip("ABC", (40, 36, 18)):
        ws.column_dimensions[col].width = w

    az_fill = PatternFill("solid", fgColor=AZ)
    gd_fill = PatternFill("solid", fgColor=GD)
    paper = PatternFill("solid", fgColor=PP)
    entry = PatternFill("solid", fgColor="FFFFFF")
    thin = Side(style="thin", color=LN)
    box = Border(left=thin, right=thin, top=thin, bottom=thin)

    def merged(row, text, font, fill=None, height=None, align="left"):
        ws.merge_cells(f"A{row}:C{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = font
        cell.alignment = Alignment(horizontal=align, vertical="center")
        if fill:
            for c in range(1, 4):
                ws.cell(row=row, column=c).fill = fill
        if height:
            ws.row_dimensions[row].height = height

    merged(1, "ORISEI FREIGHT SOLUTIONS LLC", Font(bold=True, size=16, color="FFFFFF"), az_fill, 30)
    merged(2, "MEMBER PERSONAL NET WORTH STATEMENT", Font(bold=True, size=11, color=GD), az_fill, 20)
    merged(3, "Confidential — for partnership insurance underwriting (BMC-84 surety, liability & cargo programs)",
           Font(size=8.5, color="B9C6D4"), az_fill, 16)
    merged(4, "Minneapolis–St. Paul, MN · (763) 443-4459 · oliver@oriseifreightsolutions.com",
           Font(size=8.5, color="5B6472"), paper, 14)
    merged(5, "Enter market values in the yellow VALUE column — totals and net worth calculate automatically. "
              "Save and email back to Oliver, or enter it in the Command Deck under Net Worth.",
           Font(size=8.5, italic=True, color="5B6472"), paper, 26)
    ws.cell(row=5, column=1).alignment = Alignment(wrap_text=True, vertical="center")

    r = 7
    for label, prefill in [("MEMBER NAME", member_name), ("STATEMENT DATE (AS OF)", ""), ("PHONE / EMAIL", "")]:
        ws.cell(row=r, column=1, value=label).font = Font(bold=True, size=9, color=AZ)
        ws.merge_cells(f"B{r}:C{r}")
        cell = ws.cell(row=r, column=2, value=prefill)
        cell.fill = entry
        cell.border = box
        cell.font = Font(size=10)
        r += 1
    r += 1

    val_fill = PatternFill("solid", fgColor="FFF6DA")

    def section(row, title, categories, rows_per=2):
        merged(row, title, Font(bold=True, size=11, color=AZ), gd_fill, 20)
        row += 1
        for col, head in [(1, "CATEGORY"), (2, "DESCRIPTION (what / where)"), (3, "VALUE ($)")]:
            c = ws.cell(row=row, column=col, value=head)
            c.font = Font(bold=True, size=8, color="5B6472")
            c.border = Border(bottom=thin)
        row += 1
        first_val = row
        for cat in categories:
            for i in range(rows_per):
                ws.cell(row=row, column=1, value=cat if i == 0 else "").font = Font(size=9, color="1C2430")
                dcell = ws.cell(row=row, column=2)
                vcell = ws.cell(row=row, column=3)
                dcell.fill = entry
                dcell.border = box
                dcell.font = Font(size=9)
                vcell.fill = val_fill
                vcell.border = box
                vcell.number_format = '#,##0'
                vcell.font = Font(size=10)
                row += 1
        return row, first_val, row - 1

    r, a_first, a_last = section(r, "SECTION 1 · ASSETS (what you own)", ASSET_CATEGORIES)
    ws.cell(row=r, column=1, value="TOTAL ASSETS").font = Font(bold=True, size=10)
    ta = ws.cell(row=r, column=3, value=f"=SUM(C{a_first}:C{a_last})")
    ta.font = Font(bold=True, size=11, color=AZ)
    ta.number_format = '"$"#,##0'
    total_assets_row = r
    r += 2

    r, l_first, l_last = section(r, "SECTION 2 · LIABILITIES (what you owe)", LIABILITY_CATEGORIES)
    ws.cell(row=r, column=1, value="TOTAL LIABILITIES").font = Font(bold=True, size=10)
    tl = ws.cell(row=r, column=3, value=f"=SUM(C{l_first}:C{l_last})")
    tl.font = Font(bold=True, size=11, color="8A2E1F")
    tl.number_format = '"$"#,##0'
    total_liab_row = r
    r += 1
    ws.cell(row=r, column=1, value="NET WORTH (assets − liabilities)").font = Font(bold=True, size=11)
    nw = ws.cell(row=r, column=3, value=f"=C{total_assets_row}-C{total_liab_row}")
    nw.font = Font(bold=True, size=12, color=GD)
    nw.number_format = '"$"#,##0'
    nw.fill = az_fill
    r += 2

    merged(r, "I certify that the information above is true and complete to the best of my knowledge, provided for "
              "the purpose of insurance and surety underwriting for Orisei Freight Solutions LLC.",
           Font(size=8.5, italic=True, color="1C2430"), None, 26)
    ws.cell(row=r, column=1).alignment = Alignment(wrap_text=True, vertical="center")
    r += 2
    ws.cell(row=r, column=1, value="MEMBER SIGNATURE").font = Font(bold=True, size=9, color=AZ)
    sig = ws.cell(row=r, column=2)
    sig.fill = entry
    sig.border = box
    ws.cell(row=r, column=3, value="DATE:").font = Font(bold=True, size=9, color=AZ)
    r += 1
    dt = ws.cell(row=r, column=3)
    dt.fill = entry
    dt.border = box

    for row in ws.iter_rows(min_row=1, max_row=r, min_col=1, max_col=3):
        for cell in row:
            if cell.fill.fgColor.rgb in (None, "00000000") and not cell.fill.patternType:
                cell.fill = paper
    ws.freeze_panes = "A6"
    ws.print_area = f"A1:C{r}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    _ = get_column_letter  # noqa: F841

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class SendFormIn(BaseModel):
    email: str = Field(..., min_length=3, max_length=160)


def build_net_worth_router(*, db, get_current_user: Callable) -> APIRouter:
    router = APIRouter(prefix="/net-worth", tags=["net-worth"])

    async def _seed():
        if await db.net_worth_members.count_documents({}) == 0:
            await db.net_worth_members.insert_many([
                {"id": f"NW-{uuid.uuid4().hex[:8].upper()}", "member_name": n, "as_of_date": "",
                 "assets": [], "liabilities": [], "status": "draft",
                 "created_at": _now(), "updated_at": _now()} for n in MEMBERS])

    @router.get("")
    async def overview(_=Depends(get_current_user)):
        await _seed()
        rows = await db.net_worth_members.find({}, {"_id": 0}).sort("member_name", 1).to_list(20)
        for r in rows:
            r.update(_totals(r))
        return {"members": rows,
                "asset_categories": ASSET_CATEGORIES, "liability_categories": LIABILITY_CATEGORIES,
                "combined": {"total_assets": round(sum(r["total_assets"] for r in rows), 2),
                             "total_liabilities": round(sum(r["total_liabilities"] for r in rows), 2),
                             "net_worth": round(sum(r["net_worth"] for r in rows), 2),
                             "submitted": sum(1 for r in rows if r["status"] == "submitted")}}

    @router.put("/members/{mid}")
    async def save_member(mid: str, payload: MemberStatement, _=Depends(get_current_user)):
        if payload.status not in ("draft", "submitted"):
            raise HTTPException(400, "status must be draft or submitted")
        doc = payload.model_dump()
        doc["updated_at"] = _now()
        r = await db.net_worth_members.find_one_and_update(
            {"id": mid}, {"$set": doc}, return_document=True, projection={"_id": 0})
        if not r:
            raise HTTPException(404, "Member statement not found")
        r.update(_totals(r))
        return {"ok": True, "member": r}

    @router.post("/members")
    async def add_member(payload: MemberStatement, _=Depends(get_current_user)):
        doc = {**payload.model_dump(), "id": f"NW-{uuid.uuid4().hex[:8].upper()}",
               "created_at": _now(), "updated_at": _now()}
        await db.net_worth_members.insert_one(dict(doc))
        doc.update(_totals(doc))
        return {"ok": True, "member": doc}

    @router.delete("/members/{mid}")
    async def delete_member(mid: str, _=Depends(get_current_user)):
        r = await db.net_worth_members.delete_one({"id": mid})
        if r.deleted_count == 0:
            raise HTTPException(404, "Member statement not found")
        return {"ok": True}

    @router.get("/template.pdf")
    async def template_pdf(_=Depends(get_current_user)):
        return Response(build_template_pdf(), media_type="application/pdf", headers={
            "Content-Disposition": 'attachment; filename="Orisei-Member-Net-Worth-Template.pdf"'})

    @router.get("/template.xlsx")
    async def template_xlsx(member: str = "", _=Depends(get_current_user)):
        fname = "Orisei-Member-Net-Worth-Template.xlsx"
        if member:
            fname = f"Orisei-Net-Worth-{member.replace(' ', '-')}.xlsx"
        return Response(
            build_template_xlsx(member.strip()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'})

    @router.get("/master.pdf")
    async def master_pdf(_=Depends(get_current_user)):
        await _seed()
        rows = await db.net_worth_members.find({}, {"_id": 0}).sort("member_name", 1).to_list(20)
        return Response(build_master_pdf(rows), media_type="application/pdf", headers={
            "Content-Disposition": 'attachment; filename="Orisei-Partnership-Net-Worth-Master.pdf"'})

    @router.post("/members/{mid}/send-form")
    async def send_form(mid: str, payload: SendFormIn, user=Depends(get_current_user)):
        to = payload.email.strip()
        if "@" not in to:
            raise HTTPException(400, "Valid email required")
        m = await db.net_worth_members.find_one({"id": mid}, {"_id": 0})
        if not m:
            raise HTTPException(404, "Member statement not found")
        first = m["member_name"].split(" ")[0]
        subject = f"Orisei — your net worth statement form ({m['member_name']})"
        html = f"""
        <div style="font-family:Arial,Helvetica,sans-serif;max-width:620px;margin:0 auto;color:#1a202c">
          <div style="background:#123B5C;padding:18px 26px;border-radius:8px 8px 0 0">
            <span style="color:#C9A227;font-size:18px;font-weight:800;letter-spacing:1px">ORISEI FREIGHT SOLUTIONS</span>
          </div>
          <div style="border:1px solid #e2e8f0;border-top:none;padding:24px;border-radius:0 0 8px 8px">
            <p>{first},</p>
            <p>To finish our partnership insurance application we need a personal net worth statement from
            each member. The branded form is attached — it takes about 15 minutes:</p>
            <ul style="font-size:13.5px;color:#334155">
              <li>List current market values for each asset and liability category</li>
              <li>Sign and date the certification on page 2</li>
              <li>Send it back to Oliver, or enter it directly in the Command Deck under <b>Net Worth</b></li>
            </ul>
            <p style="font-size:13px;color:#475569">Everything stays member-confidential — it rolls into one
            master partnership statement used only for underwriting.</p>
            <p>— Orisei Freight Solutions<br/>(763) 443-4459</p>
          </div>
        </div>"""
        from routes.orisei_auto_digest import _resend_creds, _send_via_resend
        creds = await _resend_creds(db)
        res = await _send_via_resend(creds, to=to, subject=subject, html=html,
                                     pdf_bytes=build_template_pdf(),
                                     pdf_filename="Orisei-Member-Net-Worth-Template.pdf",
                                     extra_attachments=[{
                                         "filename": f"Orisei-Net-Worth-{m['member_name'].replace(' ', '-')}.xlsx",
                                         "content": build_template_xlsx(m["member_name"]),
                                     }]) \
            if creds else {"sent": False, "error": "no_resend_creds"}
        status = "sent" if res.get("sent") else "recorded_no_key"
        await db.outbound_emails.insert_one({
            "to": to, "subject": subject, "html": html, "status": status, "error": res.get("error"),
            "kind": "net_worth_form", "member_id": mid, "at": _now(),
            "sent_by": getattr(user, "name", "system")})
        await db.net_worth_members.update_one(
            {"id": mid}, {"$set": {"form_sent_to": to, "form_sent_at": _now(),
                                   "form_send_status": status, "updated_at": _now()}})
        return {"ok": True, "sent": res.get("sent", False), "status": status, "to": to}

    return router

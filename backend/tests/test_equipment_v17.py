"""
Iteration 17 — Equipment / Yard Report module tests.

Covers:
  POST   /api/equipment/upload          (parse + persist .xlsx)
  GET    /api/equipment/reports         (list)
  GET    /api/equipment/reports/{id}    (single + 404)
  DELETE /api/equipment/reports/{id}    (admin-only RBAC)
  GET    /api/equipment/analytics       (snapshot + trends)
"""
import io
import os
import pytest
from datetime import date
from openpyxl import Workbook

BASE_URL = os.environ.get("REACT_APP_BACKEND_URL").rstrip("/")


def _build_yard_xlsx() -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Yard"
    ws["A1"] = date(2026, 5, 11)
    ws["B2"] = "Door"
    ws["C2"] = "Carrier"
    ws["D2"] = "Trailer Number"
    DOORS = [
        (None, 23, None, None), (None, 22, None, None), (None, 21, None, None),
        (date(2026, 5, 4), 20, "XPO", "463-8871"),
        (date(2026, 4, 29), 19, "Tennant China", "HAMU4351029"),
        (date(2026, 5, 1), 18, "Estes", "524553"),
        (None, 17, None, None),
        (date(2026, 4, 29), 16, "IPC - Soteco", "HAMU2505296"),
        (date(2026, 4, 23), 15, "Tennant China", "HAMU1802556"),
        (None, 14, "UPS / Fed EX / DHL", None),
        (None, 13, None, None),
        ("OUTBOUND", 12, "Premier A-1", "53174"),
        ("OUTBOUND", 11, "Container", "BEAU6298522"),
        ("OUTBOUND", 10, None, None), ("OUTBOUND", 9, None, None),
        ("OUTBOUND", 8, "Challenger", "953541"),
        ("OUTBOUND", 7, None, None),
        ("OUTBOUND", 6, "Saia", "486813"),
        ("OUTBOUND", 5, "Dayton", "718472"),
        ("OUTBOUND", 4, "R&L", "SF3120"),
        ("OUTBOUND", 3, "Estes", "515435"),
        ("OUTBOUND", 2, "Estes", "517536"),
        (None, 1, None, None),
    ]
    for i, (d, door, carrier, trailer) in enumerate(DOORS, start=3):
        ws.cell(row=i, column=1, value=d)
        ws.cell(row=i, column=2, value=door)
        ws.cell(row=i, column=3, value=carrier)
        ws.cell(row=i, column=4, value=trailer)

    ws["F2"] = "Loaded Trailers (Inbound)"
    LOADED_IN = [
        ("Averitt", 547036, date(2026, 5, 4)),
        ("Averitt", 542906, date(2026, 5, 5)),
        ("Estes", 538111, date(2026, 5, 4)),
        ("Estes", 527746, date(2026, 5, 5)),
        ("Holland", 31905, date(2026, 5, 8)),
        ("Saia", 488063, date(2026, 5, 7)),
        ("Meisler", 1531524, None),
    ]
    for i, (c, n, d) in enumerate(LOADED_IN, start=3):
        ws.cell(row=i, column=6, value=c); ws.cell(row=i, column=7, value=n); ws.cell(row=i, column=8, value=d)

    ws["J2"] = "Loaded Trailers (Outbound)"
    LOADED_OUT = [
        ("Averitt", 548823, "Sealed"),
        ("Averitt", 549057, "Sealed"),
        ("Averitt", 547813, "Sealed"),
        ("Averitt", 542828, "Sealed"),
        ("Container", "SEGU4963266", None),
    ]
    for i, (c, n, s) in enumerate(LOADED_OUT, start=3):
        ws.cell(row=i, column=10, value=c); ws.cell(row=i, column=11, value=n); ws.cell(row=i, column=12, value=s)

    ws["F26"] = "Empty Trailers"
    EMPTY = [("Averitt", 547220), ("Copeland", 531305), ("Estes", 524160),
             ("Holland", 31657), ("R&L", "OF2499")]
    for i, (c, n) in enumerate(EMPTY, start=27):
        ws.cell(row=i, column=6, value=c); ws.cell(row=i, column=7, value=n)

    ws["J26"] = "Empty Containers (Outbound)"
    for i, c in enumerate(["TLLU8671758", "GAOU6719522", "HAMU1284246"], start=27):
        ws.cell(row=i, column=10, value=c)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


class TestEquipmentUpload:
    def test_upload_xlsx_success(self, admin_session):
        content = _build_yard_xlsx()
        files = {"file": ("yard_test.xlsx", content, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")}
        import requests
        r = requests.post(f"{BASE_URL}/api/equipment/upload", files=files,
                          headers={"Authorization": f"Bearer {admin_session['token']}"}, timeout=30)
        assert r.status_code == 200, r.text
        d = r.json()
        assert d["report_id"].startswith("YARD-")
        assert d["report_date"] == "2026-05-11"
        assert len(d["doors"]) >= 20
        assert len(d["loaded_inbound"]) >= 5
        assert len(d["loaded_outbound"]) >= 4
        assert len(d["empty_trailers"]) >= 3
        assert len(d["empty_containers"]) >= 2
        pytest.uploaded_report_id = d["report_id"]

    def test_upload_rejects_non_xlsx(self, admin_session):
        import requests
        r = requests.post(f"{BASE_URL}/api/equipment/upload",
                          files={"file": ("bad.txt", b"hello", "text/plain")},
                          headers={"Authorization": f"Bearer {admin_session['token']}"}, timeout=15)
        assert r.status_code == 400


class TestEquipmentReports:
    def test_list_reports(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/equipment/reports")
        assert r.status_code == 200
        reports = r.json()
        assert isinstance(reports, list) and len(reports) > 0
        first = reports[0]
        for k in ["doors_total", "doors_occupied", "doors_empty",
                  "loaded_inbound", "loaded_outbound", "empty_trailers",
                  "empty_containers", "report_id", "report_date"]:
            assert k in first, f"missing key: {k}"

    def test_get_report_by_id(self, admin_client):
        rid = getattr(pytest, "uploaded_report_id", None)
        assert rid
        r = admin_client.get(f"{BASE_URL}/api/equipment/reports/{rid}")
        assert r.status_code == 200
        d = r.json()
        assert d["report_id"] == rid
        assert "doors" in d and "loaded_inbound" in d

    def test_get_report_404(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/equipment/reports/YARD-DOESNOTEXIST")
        assert r.status_code == 404


class TestAnalytics:
    def test_analytics_has_all_fields(self, admin_client):
        r = admin_client.get(f"{BASE_URL}/api/equipment/analytics")
        assert r.status_code == 200
        a = r.json()
        for k in ["snapshot", "trend", "carrier_mix", "dwell",
                  "door_occupancy_pct", "sealed_pct", "report_count"]:
            assert k in a, f"missing analytics key: {k}"
        assert a["report_count"] >= 1
        snap = a["snapshot"]
        assert snap is not None
        for k in ["total_on_site", "sealed_count", "sealed_pct",
                  "door_occupancy_pct", "doors_total", "doors_occupied"]:
            assert k in snap, f"missing snapshot key: {k}"
        assert isinstance(a["carrier_mix"], list)
        assert isinstance(a["trend"], list)
        assert isinstance(a["dwell"], list)


class TestRbacDelete:
    def test_delete_forbidden_for_dispatcher(self, dispatcher_client):
        rid = getattr(pytest, "uploaded_report_id", None)
        assert rid
        r = dispatcher_client.delete(f"{BASE_URL}/api/equipment/reports/{rid}")
        assert r.status_code == 403, f"expected 403 got {r.status_code}"

    def test_delete_forbidden_for_auditor(self, auditor_client):
        rid = getattr(pytest, "uploaded_report_id", None)
        assert rid
        r = auditor_client.delete(f"{BASE_URL}/api/equipment/reports/{rid}")
        assert r.status_code == 403

    def test_delete_admin_ok(self, admin_client):
        rid = getattr(pytest, "uploaded_report_id", None)
        assert rid
        r = admin_client.delete(f"{BASE_URL}/api/equipment/reports/{rid}")
        assert r.status_code == 200
        # verify gone
        g = admin_client.get(f"{BASE_URL}/api/equipment/reports/{rid}")
        assert g.status_code == 404

"""
Equipment / Yard report module.

Parses a daily yard snapshot Excel sheet (Door assignments, Loaded Inbound,
Loaded Outbound, Empty Trailers, Empty Containers) and exposes:

  POST   /api/equipment/upload            — upload .xlsx, parse, persist
  GET    /api/equipment/reports           — list reports (id, date, counts)
  GET    /api/equipment/reports/{id}      — full report payload
  DELETE /api/equipment/reports/{id}      — remove a report
  GET    /api/equipment/analytics         — cross-report analytics for the dashboard

The parser is intentionally tolerant — it scans the workbook for section
header strings ("Loaded Trailers (Inbound)", "Empty Trailers", "Door", ...)
and reads the rows beneath until a blank row or the next section is reached.
This lets the same parser handle small column drift between reports.
"""
from __future__ import annotations

import io
import re
import uuid
from collections import Counter
from datetime import datetime, timezone, date
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, Depends, File, HTTPException, UploadFile
from openpyxl import load_workbook
from pydantic import BaseModel


router = APIRouter(prefix="/equipment", tags=["equipment"])


# -------------------- Models --------------------
class DoorRow(BaseModel):
    door: int
    carrier: Optional[str] = None
    trailer_no: Optional[str] = None
    status: Optional[str] = None  # OUTBOUND / loaded date / None
    date: Optional[str] = None    # ISO date if shown


class TrailerRow(BaseModel):
    carrier: Optional[str] = None
    trailer_no: Optional[str] = None
    date: Optional[str] = None
    status: Optional[str] = None  # e.g. "Sealed"


class EquipmentReport(BaseModel):
    report_id: str
    report_date: str  # ISO date
    uploaded_at: str
    uploaded_by: str
    filename: str
    doors: List[DoorRow] = []
    loaded_inbound: List[TrailerRow] = []
    loaded_outbound: List[TrailerRow] = []
    empty_trailers: List[TrailerRow] = []
    empty_containers: List[TrailerRow] = []


# -------------------- Parser helpers --------------------
SECTION_PATTERNS = {
    "loaded_inbound": re.compile(r"loaded\s+trailers?\s*\(inbound\)", re.I),
    "loaded_outbound": re.compile(r"loaded\s+trailers?\s*\(outbound\)", re.I),
    "empty_trailers": re.compile(r"empty\s+trailers", re.I),
    "empty_containers": re.compile(r"empty\s+containers", re.I),
    "doors_header": re.compile(r"^\s*door\s*$", re.I),
}


def _coerce_str(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    s = str(v).strip()
    return s or None


def _coerce_date(v: Any) -> Optional[str]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    s = str(v).strip()
    if not s:
        return None
    # Try common date formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d/%m/%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _coerce_int(v: Any) -> Optional[int]:
    try:
        if v is None or v == "":
            return None
        return int(float(v))
    except (TypeError, ValueError):
        return None


def _scan_section_start(ws, label_regex: re.Pattern) -> Optional[Tuple[int, int]]:
    """Return (row, col) of the first cell whose value matches label_regex."""
    for row in ws.iter_rows(values_only=False):
        for cell in row:
            v = cell.value
            if isinstance(v, str) and label_regex.search(v):
                return cell.row, cell.column
    return None


def _read_trailer_section(ws, start_row: int, start_col: int, max_rows: int = 60) -> List[TrailerRow]:
    """Read a trailer table starting at (start_row+1, start_col).
    Expected columns: Carrier | Trailer # | (Date or Status).
    Stops at first row where both carrier and trailer# are empty.
    """
    out: List[TrailerRow] = []
    for i in range(1, max_rows + 1):
        r = start_row + i
        a = ws.cell(row=r, column=start_col).value
        b = ws.cell(row=r, column=start_col + 1).value
        c = ws.cell(row=r, column=start_col + 2).value
        carrier = _coerce_str(a)
        trailer = _coerce_str(b)
        if not carrier and not trailer:
            # Allow one blank row tolerance, then stop
            a2 = ws.cell(row=r + 1, column=start_col).value
            b2 = ws.cell(row=r + 1, column=start_col + 1).value
            if not _coerce_str(a2) and not _coerce_str(b2):
                break
            continue
        # Third column: try date first, then plain string (e.g. "Sealed")
        c_date = _coerce_date(c)
        c_status = _coerce_str(c) if not c_date else None
        out.append(TrailerRow(carrier=carrier, trailer_no=trailer, date=c_date, status=c_status))
    return out


def _read_doors(ws, header_row: int, header_col: int, max_rows: int = 30) -> List[DoorRow]:
    """Reads the door assignment table. Header is at (header_row, header_col) on the
    'Door' label. We assume columns: [date?, door#, carrier, trailer_no].
    Some doors have a date in the leftmost column (e.g. last-touched date).
    """
    # Locate optional date column to the LEFT of Door
    date_col = header_col - 1 if header_col >= 2 else None
    door_col = header_col
    carrier_col = header_col + 1
    trailer_col = header_col + 2
    out: List[DoorRow] = []
    for i in range(1, max_rows + 1):
        r = header_row + i
        door_val = _coerce_int(ws.cell(row=r, column=door_col).value)
        date_val = _coerce_str(ws.cell(row=r, column=date_col).value) if date_col else None
        carrier = _coerce_str(ws.cell(row=r, column=carrier_col).value)
        trailer = _coerce_str(ws.cell(row=r, column=trailer_col).value)
        if door_val is None and not carrier and not trailer:
            # Stop only after a clean blank row (door# is the most reliable signal)
            if ws.cell(row=r + 1, column=door_col).value is None:
                break
            continue
        # Detect "OUTBOUND" placeholder in the date column
        status = None
        date_iso = None
        if date_val:
            if "OUTBOUND" in date_val.upper():
                status = "OUTBOUND"
            else:
                date_iso = _coerce_date(date_val)
        out.append(DoorRow(
            door=door_val if door_val is not None else 0,
            carrier=carrier, trailer_no=trailer,
            status=status, date=date_iso,
        ))
    return out


def parse_yard_xlsx(content: bytes) -> Dict[str, Any]:
    """Top-level parser. Returns a dict ready to merge into an EquipmentReport."""
    wb = load_workbook(io.BytesIO(content), data_only=True, read_only=False)
    ws = wb.active

    # Report date: scan top-left cells for the first date value
    report_date_iso: Optional[str] = None
    for row in ws.iter_rows(min_row=1, max_row=5, values_only=False):
        for cell in row:
            d = _coerce_date(cell.value)
            if d:
                report_date_iso = d
                break
        if report_date_iso:
            break
    if not report_date_iso:
        report_date_iso = datetime.now(timezone.utc).date().isoformat()

    # Doors
    doors: List[DoorRow] = []
    door_header = _scan_section_start(ws, SECTION_PATTERNS["doors_header"])
    if door_header:
        doors = _read_doors(ws, door_header[0], door_header[1])

    # Trailer / container sections
    sections: Dict[str, List[TrailerRow]] = {
        "loaded_inbound": [], "loaded_outbound": [],
        "empty_trailers": [], "empty_containers": [],
    }
    for key in sections:
        pos = _scan_section_start(ws, SECTION_PATTERNS[key])
        if pos:
            # The header row labels are typically on the row below the section
            # title ("Carrier / Trailer # / Date"); but data also starts there
            # in this template. Skip one row past the section title and read.
            sections[key] = _read_trailer_section(ws, pos[0] + 1, pos[1])

    return {
        "report_date": report_date_iso,
        "doors": [d.model_dump() for d in doors],
        "loaded_inbound": [r.model_dump() for r in sections["loaded_inbound"]],
        "loaded_outbound": [r.model_dump() for r in sections["loaded_outbound"]],
        "empty_trailers": [r.model_dump() for r in sections["empty_trailers"]],
        "empty_containers": [r.model_dump() for r in sections["empty_containers"]],
    }


# -------------------- Endpoints --------------------
def _summary(report: Dict[str, Any]) -> Dict[str, Any]:
    occupied = sum(1 for d in report.get("doors", []) if (d.get("carrier") or d.get("trailer_no")))
    total_doors = len(report.get("doors", []))
    return {
        "doors_total": total_doors,
        "doors_occupied": occupied,
        "doors_empty": total_doors - occupied,
        "loaded_inbound": len(report.get("loaded_inbound", [])),
        "loaded_outbound": len(report.get("loaded_outbound", [])),
        "empty_trailers": len(report.get("empty_trailers", [])),
        "empty_containers": len(report.get("empty_containers", [])),
    }


def register_equipment_routes(api_router, db, get_current_user, require_role):
    """Bound at server startup. Receives the parent router + DB + auth deps."""

    @router.post("/upload", response_model=EquipmentReport)
    async def upload_report(
        file: UploadFile = File(...),
        user=Depends(require_role("admin", "auditor", "dispatcher")),
    ):
        if not (file.filename or "").lower().endswith((".xlsx", ".xlsm")):
            raise HTTPException(status_code=400, detail="Upload an .xlsx file")
        content = await file.read()
        if not content:
            raise HTTPException(status_code=400, detail="Empty file")
        try:
            parsed = parse_yard_xlsx(content)
        except Exception as e:
            raise HTTPException(status_code=400, detail=f"Could not parse Excel: {e}")
        doc = {
            "report_id": f"YARD-{uuid.uuid4().hex[:8].upper()}",
            "report_date": parsed["report_date"],
            "uploaded_at": datetime.now(timezone.utc).isoformat(),
            "uploaded_by": user.name,
            "filename": file.filename,
            **parsed,
        }
        await db.equipment_reports.insert_one(dict(doc))
        return EquipmentReport(**doc)

    @router.get("/reports")
    async def list_reports(_=Depends(get_current_user)):
        docs = await db.equipment_reports.find({}, {"_id": 0}).sort("report_date", -1).limit(200).to_list(200)
        return [{**_summary(d),
                 "report_id": d["report_id"],
                 "report_date": d["report_date"],
                 "uploaded_at": d["uploaded_at"],
                 "uploaded_by": d["uploaded_by"],
                 "filename": d["filename"]} for d in docs]

    @router.get("/reports/{report_id}", response_model=EquipmentReport)
    async def get_report(report_id: str, _=Depends(get_current_user)):
        doc = await db.equipment_reports.find_one({"report_id": report_id}, {"_id": 0})
        if not doc:
            raise HTTPException(status_code=404, detail="Report not found")
        return EquipmentReport(**doc)

    @router.delete("/reports/{report_id}")
    async def delete_report(report_id: str, _=Depends(require_role("admin"))):
        r = await db.equipment_reports.delete_one({"report_id": report_id})
        if r.deleted_count == 0:
            raise HTTPException(status_code=404, detail="Report not found")
        return {"ok": True}

    @router.get("/analytics")
    async def analytics(_=Depends(get_current_user)):
        """Aggregates the most recent report (snapshot view) AND historical
        trends across all uploaded reports (timeline view)."""
        docs = await db.equipment_reports.find({}, {"_id": 0}).sort("report_date", -1).limit(180).to_list(180)
        if not docs:
            return {
                "snapshot": None,
                "trend": [],
                "carrier_mix": [],
                "dwell": [],
                "door_occupancy_pct": 0,
                "sealed_pct": 0,
                "report_count": 0,
            }
        latest = docs[0]
        # ---- Snapshot KPIs (latest report) ----
        summary = _summary(latest)
        total = summary["doors_total"] or 1
        door_occupancy_pct = round(100 * summary["doors_occupied"] / total, 1)
        outbound = latest.get("loaded_outbound", [])
        sealed = sum(1 for r in outbound if (r.get("status") or "").lower() == "sealed")
        sealed_pct = round(100 * sealed / len(outbound), 1) if outbound else 0

        # ---- Carrier mix across ALL trailers in the latest snapshot ----
        all_rows = (
            latest.get("loaded_inbound", []) +
            latest.get("loaded_outbound", []) +
            latest.get("empty_trailers", []) +
            [d for d in latest.get("doors", []) if d.get("carrier")]
        )
        cmix = Counter([(r.get("carrier") or "Unknown").strip() for r in all_rows if r.get("carrier")])
        carrier_mix = [{"carrier": k, "count": v} for k, v in cmix.most_common(15)]

        # ---- Dwell time on loaded inbound (days since arrival vs report_date) ----
        try:
            report_d = datetime.strptime(latest["report_date"], "%Y-%m-%d").date()
        except Exception:
            report_d = datetime.now(timezone.utc).date()
        dwell: List[Dict[str, Any]] = []
        for r in latest.get("loaded_inbound", []):
            d = r.get("date")
            if not d:
                continue
            try:
                arrived = datetime.strptime(d, "%Y-%m-%d").date()
                days = (report_d - arrived).days
                dwell.append({
                    "carrier": r.get("carrier"),
                    "trailer_no": r.get("trailer_no"),
                    "arrived": d,
                    "days": days,
                    "bucket": "0-1" if days <= 1 else "2-3" if days <= 3 else "4-7" if days <= 7 else "8+",
                })
            except Exception:
                continue

        # ---- Trend across reports (newest last for chart) ----
        trend = []
        for d in reversed(docs):
            s = _summary(d)
            trend.append({
                "date": d["report_date"],
                "loaded_inbound": s["loaded_inbound"],
                "loaded_outbound": s["loaded_outbound"],
                "empty_trailers": s["empty_trailers"],
                "empty_containers": s["empty_containers"],
                "doors_occupied": s["doors_occupied"],
            })

        return {
            "snapshot": {
                **summary,
                "report_id": latest["report_id"],
                "report_date": latest["report_date"],
                "uploaded_at": latest["uploaded_at"],
                "door_occupancy_pct": door_occupancy_pct,
                "sealed_pct": sealed_pct,
                "sealed_count": sealed,
                "total_on_site": (
                    summary["doors_occupied"]
                    + summary["loaded_inbound"]
                    + summary["loaded_outbound"]
                    + summary["empty_trailers"]
                    + summary["empty_containers"]
                ),
            },
            "trend": trend,
            "carrier_mix": carrier_mix,
            "dwell": dwell,
            "door_occupancy_pct": door_occupancy_pct,
            "sealed_pct": sealed_pct,
            "report_count": len(docs),
        }

    api_router.include_router(router)
